# NHG Weather Data Pipeline — app.py
# Imports
import streamlit as st
import pandas as pd
import os
import json
import numpy as np
import csv
import io
import re
import matplotlib.pyplot as plt
from datetime import datetime, timedelta, timezone
from suntime import Sun
from pypdf import PdfReader

try:
    import openpyxl  # needed for MetadataLog parsing
except ImportError:
    openpyxl = None

# Page Config
st.set_page_config(page_title="NHG Weather Pipeline", layout="wide")

# --- File paths for persisted config ---
MAPPING_FILE = "column_mapping.json"
GROUPS_FILE = "instrument_groups.json"
STATION_CONFIG_FILE = "station_configs.json"

# ---------------------------------------------------------------------------
# SENSOR_THRESHOLDS — two-tier threshold table derived from RefSensorThresholds.xlsx
#
# Each entry maps a canonical column name to a dict with:
#   r_min / r_max  — hard physical limits  → flag 'R' if breached
#   c_min / c_max  — soft caution limits   → flag 'C' if breached (but within R)
#   None means "no limit for this tier" — that check is skipped entirely.
#
# Special string values (e.g. 'H-50', 'SWin_Avg') are resolved at runtime.
# ---------------------------------------------------------------------------
SENSOR_THRESHOLDS = {
    # --- CR350 Data Logger ---
    'BattV_Avg':      {'r_min': +9.6,  'r_max': 19.0,   'c_min': 10.0,  'c_max': 16.0},
    'RECORD':         {'r_min': 0,     'r_max': None,   'c_min': None,  'c_max': None},
    'PTemp_C_Avg':    {'r_min': -40.0, 'r_max': 70.0,   'c_min': None,  'c_max': None},

    # --- Apogee ST-110 Ground Surface Temperature Probe ---
    'Stmp_Avg':       {'r_min': -40.0, 'r_max': 70.0,   'c_min': None,  'c_max': None},
    # --- Campbell Scientific Model-109 Ground Temperature Probe ---
    'Gtmp_Avg':       {'r_min': -50.0, 'r_max': 70.0,   'c_min': None,  'c_max': None},
    # --- Campbell Scientific T109 Air Temperature Probe (standalone) ---
    'T109_AirT_C_Avg': {'r_min': -50.0, 'r_max': 70.0,  'c_min': None,  'c_max': None},

    # --- Campbell Scientific ClimaVUE50 Compact Weather Sensor ---
    'AirT_C_Avg':     {'r_min': -50.0, 'r_max': 60.0,   'c_min': None,  'c_max': None},
    'RHT_C_Avg':      {'r_min': -40.0, 'r_max': 60.0,   'c_min': None,  'c_max': None},
    'RHT_Avg':        {'r_min': -40.0, 'r_max': 60.0,   'c_min': None,  'c_max': None},
    'SlrFD_W_Avg':    {'r_min': 0.0,   'r_max': 1750.0, 'c_min': 0.0,   'c_max': 1360.0},
    'Rain_mm_Tot':    {'r_min': 0.0,   'r_max': 100.0,  'c_min': 0.0,   'c_max': 12.5},
    'Strikes_Tot':    {'r_min': 0.0,   'r_max': 65535.0,'c_min': None,  'c_max': None},
    'Dist_km_Avg':    {'r_min': 0.0,   'r_max': 40.0,   'c_min': None,  'c_max': None},
    'WS_ms_Avg':      {'r_min': 0.0,   'r_max': 30.0,   'c_min': None,  'c_max': None},
    'WindDir':        {'r_min': 0.0,   'r_max': 359.0,  'c_min': None,  'c_max': None},
    'MaxWS_ms':       {'r_min': 0.0,   'r_max': 30.0,   'c_min': None,  'c_max': None},
    # Alias
    'MaxWS_ms_Avg':   {'r_min': 0.0,   'r_max': 30.0,   'c_min': None,  'c_max': None},
    'VP_hPa_Avg':     {'r_min': 0.0,   'r_max': 47.0,   'c_min': None,  'c_max': None},
    'RH':             {'r_min': 0.0,   'r_max': 100.0,  'c_min': None,  'c_max': None},
    'BP_hPa_Avg':     {'r_min': 500.0, 'r_max': 1100.0, 'c_min': None,  'c_max': None},
    'TiltNS_deg_Avg': {'r_min': -90.0, 'r_max': 90.0,   'c_min': -3.0,  'c_max': 3.0},
    'TiltWE_deg_Avg': {'r_min': -90.0, 'r_max': 90.0,   'c_min': -3.0,  'c_max': 3.0},
    'SlrTF_MJ_Tot':   {'r_min': 0.0,   'r_max': 1.575,  'c_min': 0.0,   'c_max': 1.215},

    # --- Campbell Scientific SR50 Sonic Ranger ---
    # DT hard limits from RefSensorThresholds.xlsx
    'DT_Avg':         {'r_min': 50.0,  'r_max': 1000.0, 'c_min': None,  'c_max': None},
    'Q_Avg':          {'r_min': 162.0, 'r_max': 600.0,  'c_min': None,  'c_max': 210.0},
    'TCDT_Avg':       {'r_min': 50.0,  'r_max': 1000.0, 'c_min': None,  'c_max': None},
    # DBTCDT max = H-50 (sensor-height-dependent)
    'DBTCDT_Avg':     {'r_min': 0.0,   'r_max': 'H-50', 'c_min': None,  'c_max': None},

    # --- Apogee SN-500 Net Radiometer ---
    'SWin_Avg':       {'r_min': -10.0, 'r_max': 2000.0, 'c_min': 0.0,   'c_max': 1360.0},
    'SWout_Avg':      {'r_min': -10.0, 'r_max': 2000.0, 'c_min': 0.0,   'c_max': 1360.0},
    # LW has no hard R limits per the sheet
    'LWin_Avg':       {'r_min': None,  'r_max': None,   'c_min': 0.0,   'c_max': 600.0},
    'LWout_Avg':      {'r_min': None,  'r_max': None,   'c_min': 0.0,   'c_max': 700.0},
    'SWnet_Avg':      {'r_min': -2000.0,'r_max': 2000.0,'c_min': 0.0,   'c_max': 1360.0},
    'LWnet_Avg':      {'r_min': -200.0,'r_max': 200.0,  'c_min': -150.0,'c_max': 100.0},
    'SWalbedo_Avg':   {'r_min': 0.0,   'r_max': 1.0,    'c_min': 0.05,  'c_max': 0.95},
    'NR_Avg':         {'r_min': -2200.0,'r_max': 2200.0,'c_min': -200.0,'c_max': 1000.0},

    # --- Timestamp-type columns (no numeric thresholds — Date/Time values) ---
    'MaxWS_ms_TMx':   {'r_min': None, 'r_max': None, 'c_min': None, 'c_max': None},
    'Dist_km_TMx':    {'r_min': None, 'r_max': None, 'c_min': None, 'c_max': None},

    # --- Campbell Scientific SnowVue10 Smart Snow Depth Sensor ---
    'SV_DT_Avg':      {'r_min': 40.0,  'r_max': 1000.0, 'c_min': None,  'c_max': None},
    'SV_Q_Avg':       {'r_min': 0.0,   'r_max': 600.0,  'c_min': 100.0, 'c_max': 600.0},
    'SV_TCDT_Avg':    {'r_min': 40.0,  'r_max': 1000.0, 'c_min': None,  'c_max': None},
    'SV_DBTCDT_Avg':  {'r_min': 0.0,   'r_max': 'H-40', 'c_min': None,  'c_max': None},
    'SV_Tmp_Avg':     {'r_min': -45.0, 'r_max': 50.0,   'c_min': None,  'c_max': None},
    'SV_RH':          {'r_min': None,  'r_max': None,   'c_min': 0.0,   'c_max': 70.0},
    'SV_Pitch':       {'r_min': None,  'r_max': None,   'c_min': 0.0,   'c_max': 10.0},
    'SV_Roll':        {'r_min': None,  'r_max': None,   'c_min': 0.0,   'c_max': 10.0},
    'SV_Volt':        {'r_min': None,  'r_max': None,   'c_min': 5.5,   'c_max': 24.0},
    'SV_Freq':        {'r_min': None,  'r_max': None,   'c_min': 40.0,  'c_max': 60.0},
    'SV_Alert':       {'r_min': None,  'r_max': None,   'c_min': None,  'c_max': None},
}

# ---------------------------------------------------------------------------
# DEFAULT_THRESHOLDS — kept for backward compatibility with the instrument-group
# editor UI (which still uses min/max pairs). Derived from SENSOR_THRESHOLDS r_min/r_max.
# ---------------------------------------------------------------------------
DEFAULT_THRESHOLDS = {
    col: (v['r_min'] if v['r_min'] is not None else -np.inf,
          v['r_max'] if v['r_max'] is not None and not isinstance(v['r_max'], str) else np.inf)
    for col, v in SENSOR_THRESHOLDS.items()
}

# ---------------------------------------------------------------------------
# INITIAL_INSTRUMENT_GROUPS — full R/C threshold structure per sensor group.
# Matches RefSensorThresholds.xlsx exactly.  Each column has:
#   r_min / r_max  — hard physical limits  (flag R)
#   c_min / c_max  — soft caution limits   (flag C)
# None means "no limit for this tier" — the check is skipped.
# ---------------------------------------------------------------------------
INITIAL_INSTRUMENT_GROUPS = {
    'ClimaVue50': {
        'sensor_height': 180,
        'thresholds': {
            'AirT_C_Avg':     {'r_min': -50,   'r_max': 60,    'c_min': None,  'c_max': None},
            'RHT_Avg':        {'r_min': -40,   'r_max': 60,    'c_min': None,  'c_max': None},
            'RHT_C_Avg':      {'r_min': -40,   'r_max': 60,    'c_min': None,  'c_max': None},
            'SlrFD_W_Avg':    {'r_min': 0,     'r_max': 1750,  'c_min': 0,     'c_max': 1360},
            'Rain_mm_Tot':    {'r_min': 0,     'r_max': 100,   'c_min': 0,     'c_max': 12.5},
            'Strikes_Tot':    {'r_min': 0,     'r_max': 65535, 'c_min': None,  'c_max': None},
            'Dist_km_Avg':    {'r_min': 0,     'r_max': 40,    'c_min': None,  'c_max': None},
            'WS_ms_Avg':      {'r_min': 0,     'r_max': 30,    'c_min': None,  'c_max': None},
            'WindDir':        {'r_min': 0,     'r_max': 359,   'c_min': None,  'c_max': None},
            'MaxWS_ms':       {'r_min': 0,     'r_max': 30,    'c_min': None,  'c_max': None},
            'VP_hPa_Avg':     {'r_min': 0,     'r_max': 47,    'c_min': None,  'c_max': None},
            'RH':             {'r_min': 0,     'r_max': 100,   'c_min': None,  'c_max': None},
            'BP_hPa_Avg':     {'r_min': 500,   'r_max': 1100,  'c_min': None,  'c_max': None},
            'TiltNS_deg_Avg': {'r_min': -90,   'r_max': 90,    'c_min': -3,    'c_max': 3},
            'TiltWE_deg_Avg': {'r_min': -90,   'r_max': 90,    'c_min': -3,    'c_max': 3},
            'SlrTF_MJ_Tot':   {'r_min': 0,     'r_max': 1.575, 'c_min': 0,     'c_max': 1.215},
        },
    },
    'SR50': {
        'sensor_height': 200,
        'thresholds': {
            'DT_Avg':      {'r_min': 50,  'r_max': 1000, 'c_min': None, 'c_max': None},
            'Q_Avg':       {'r_min': 162, 'r_max': 600,  'c_min': None, 'c_max': 210},
            'TCDT_Avg':    {'r_min': 50,  'r_max': 1000, 'c_min': None, 'c_max': None},
            'DBTCDT_Avg':  {'r_min': 0,   'r_max': 150,  'c_min': None, 'c_max': None},
            # Note: DBTCDT r_max = sensor_height (200) - 50 = 150
        },
    },
    'NetRadiometer': {
        'sensor_height': 180,
        'thresholds': {
            'SWin_Avg':     {'r_min': -10,   'r_max': 2000, 'c_min': 0,    'c_max': 1360},
            'SWout_Avg':    {'r_min': -10,   'r_max': 2000, 'c_min': 0,    'c_max': 1360},
            'LWin_Avg':     {'r_min': None,  'r_max': None, 'c_min': 0,    'c_max': 600},
            'LWout_Avg':    {'r_min': None,  'r_max': None, 'c_min': 0,    'c_max': 700},
            'SWnet_Avg':    {'r_min': -2000, 'r_max': 2000, 'c_min': 0,    'c_max': 1360},
            'LWnet_Avg':    {'r_min': -200,  'r_max': 200,  'c_min': -150, 'c_max': 100},
            'SWalbedo_Avg': {'r_min': 0,     'r_max': 1,    'c_min': 0.05, 'c_max': 0.95},
            'NR_Avg':       {'r_min': -2200, 'r_max': 2200, 'c_min': -200, 'c_max': 1000},
            'Stmp_Avg':     {'r_min': -40,   'r_max': 70,   'c_min': None, 'c_max': None},
            'Gtmp_Avg':     {'r_min': -50,   'r_max': 70,   'c_min': None, 'c_max': None},
        },
    },
    'System': {
        'thresholds': {
            'BattV_Avg':   {'r_min': 9.6, 'r_max': 19, 'c_min': 10,   'c_max': 16},
            'PTemp_C_Avg': {'r_min': -40,  'r_max': 70, 'c_min': None, 'c_max': None},
            'Ptmp_C_Avg':  {'r_min': -40,  'r_max': 70, 'c_min': None, 'c_max': None},
        },
    },
    'T109': {
        'sensor_height': 30,
        'thresholds': {
            # T109 air temperature is intentionally separate from ClimaVue AirT_C_Avg
            # so ClimaVue dependency rules do not apply.
            'T109_AirT_C_Avg': {'r_min': -50, 'r_max': 70, 'c_min': None, 'c_max': None},
            'Gtmp_Avg':        {'r_min': -50, 'r_max': 70, 'c_min': None, 'c_max': None},
        },
    },
    'SnowVue10': {
        'sensor_height': 200,
        'thresholds': {
            'SV_DT_Avg':     {'r_min': 40,   'r_max': 1000, 'c_min': None,  'c_max': None},
            'SV_Q_Avg':      {'r_min': 0,    'r_max': 600,  'c_min': 100,   'c_max': 600},
            'SV_TCDT_Avg':   {'r_min': 40,   'r_max': 1000, 'c_min': None,  'c_max': None},
            'SV_DBTCDT_Avg': {'r_min': 0,    'r_max': 'H-40','c_min': None, 'c_max': None},
            'SV_Tmp_Avg':    {'r_min': -45,  'r_max': 50,   'c_min': None,  'c_max': None},
            'SV_RH':         {'r_min': None, 'r_max': None, 'c_min': 0,     'c_max': 70},
            'SV_Pitch':      {'r_min': None, 'r_max': None, 'c_min': 0,     'c_max': 10},
            'SV_Roll':       {'r_min': None, 'r_max': None, 'c_min': 0,     'c_max': 10},
            'SV_Volt':       {'r_min': None, 'r_max': None, 'c_min': 5.5,   'c_max': 24},
            'SV_Freq':       {'r_min': None, 'r_max': None, 'c_min': 40,    'c_max': 60},
            'SV_Alert':      {'r_min': None, 'r_max': None, 'c_min': None,  'c_max': None},
        },
    },
}

# ---------------------------------------------------------------------------
# DEPENDENCY_CONFIG — propagates flags from source columns to dependent targets.
# Derived directly from the 'Flags_Depend' and 'Notes' columns of RefSensorThresholds.xlsx.
#
# Flag meanings (from FlagLibrary sheet):
#   R   — hard physical limit breach
#   C   — soft caution limit breach
#   E   — sensor-specific error (e.g. -9999 value)
#   DF  — one or more dependent variables is flagged R, E, or DF
#   DC  — cautionary dependency (dependent variable is flagged C or T)
#   T   — sensor tilt exceeds accuracy range (C-range tilt breach, NOT the same as R)
#   BV  — battery voltage flagged R (propagates to all sensor columns)
#   PT  — panel temperature flagged R (propagates to all sensor columns)
#   LR  — logger restart (RECORD sequence restart, propagates to all sensor columns)
#   NV  — no valid value (e.g. wind direction when wind speed == 0)
#   Z   — value < 0 at night
#   SF  — snow-free period (Sep 1 – Jun 30)
#   DZ  — divide-by-zero (e.g. albedo when SWin < 20 W/m²)
# ---------------------------------------------------------------------------
DEPENDENCY_CONFIG = [
    # -----------------------------------------------------------------------
    # System-level propagation: BV (battery voltage R) → all sensor columns
    # -----------------------------------------------------------------------
    # BV is applied programmatically in run_qc_pipeline, not listed per-column here,
    # because it affects every column. See the BV/PT/LR propagation block in the pipeline.

    # -----------------------------------------------------------------------
    # ClimaVUE50 — tilt affects solar flux, rain, and derived columns
    # -----------------------------------------------------------------------
    # Tilt R (> |90°|) → solar flux gets DC (sensor knocked over) 
    {'target': 'SlrFD_W_Avg',  'sources': ['TiltNS_deg_Avg', 'TiltWE_deg_Avg'], 'trigger_flags': ['R'], 'set_flag': 'DC'},
    # Tilt T or C (> |3°| but < |90°|) → solar flux gets T flag (tilt exceeds accuracy)
    {'target': 'SlrFD_W_Avg',  'sources': ['TiltNS_deg_Avg', 'TiltWE_deg_Avg'], 'trigger_flags': ['C'], 'set_flag': 'T'},
    # Tilt R → rain gets DC
    {'target': 'Rain_mm_Tot',  'sources': ['TiltNS_deg_Avg', 'TiltWE_deg_Avg'], 'trigger_flags': ['R'], 'set_flag': 'DC'},
    # Tilt T or C → rain gets T
    {'target': 'Rain_mm_Tot',  'sources': ['TiltNS_deg_Avg', 'TiltWE_deg_Avg'], 'trigger_flags': ['C'], 'set_flag': 'T'},

    # ClimaVUE50 — solar flux affects air temperature (energy balance correction)
    # AirT gets DC if SlrFD_W_Avg is DC or T (per Notes: "flag DC if SlrFD_W_Avg == DC or T")
    {'target': 'AirT_C_Avg',   'sources': ['SlrFD_W_Avg'],                      'trigger_flags': ['DC', 'T'], 'set_flag': 'DC'},
    # AirT gets DF if SlrFD_W_Avg or WS_ms_Avg is R, E, or DF
    {'target': 'AirT_C_Avg',   'sources': ['SlrFD_W_Avg', 'WS_ms_Avg'],         'trigger_flags': ['R', 'E', 'M', 'DF'], 'set_flag': 'DF'},

    # ClimaVUE50 — RH probe temperature affects vapour pressure
    {'target': 'VP_hPa_Avg',   'sources': ['RHT_Avg'],                          'trigger_flags': ['R', 'E', 'M', 'DF'], 'set_flag': 'DF'},

    # ClimaVUE50 — VP and AirT affect RH
    {'target': 'RH',           'sources': ['VP_hPa_Avg', 'AirT_C_Avg'],         'trigger_flags': ['R', 'E', 'M', 'DF'], 'set_flag': 'DF'},
    # RH gets DC if AirT_C_Avg is DC (per Notes: "Flag DC if AirT_C_Avg == DC")
    {'target': 'RH',           'sources': ['AirT_C_Avg'],                       'trigger_flags': ['DC'], 'set_flag': 'DC'},

    # ClimaVUE50 — solar flux affects total solar flux (SlrTF_MJ_Tot)
    # SlrTF gets DC if SlrFD_W_Avg is C, T, or DC (per Notes: "DC flag if SlrFD_W_Avg == C, T, DC")
    {'target': 'SlrTF_MJ_Tot', 'sources': ['SlrFD_W_Avg'],                      'trigger_flags': ['C', 'T', 'DC'], 'set_flag': 'DC'},
    # SlrTF gets DF if SlrFD_W_Avg is R or E
    {'target': 'SlrTF_MJ_Tot', 'sources': ['SlrFD_W_Avg'],                      'trigger_flags': ['R', 'E', 'M'], 'set_flag': 'DF'},
    # SlrTF inherits Z from SlrFD_W_Avg
    {'target': 'SlrTF_MJ_Tot', 'sources': ['SlrFD_W_Avg'],                      'trigger_flags': ['Z'], 'set_flag': 'Z'},

    # ClimaVUE50 — wind direction and gust invalid when wind speed == 0 (NV flag applied in pipeline)
    # WindDir/MaxWS also receive DF when WS has hard/error dependency failure.
    {'target': 'WindDir',      'sources': ['WS_ms_Avg'],                         'trigger_flags': ['R', 'E', 'DF', 'M'], 'set_flag': 'DF'},
    {'target': 'WindDir',      'sources': ['WS_ms_Avg'],                         'trigger_flags': ['NV'], 'set_flag': 'NV'},
    {'target': 'MaxWS_ms',     'sources': ['WS_ms_Avg'],                         'trigger_flags': ['R', 'E', 'DF', 'M'], 'set_flag': 'DF'},
    {'target': 'MaxWS_ms',     'sources': ['WS_ms_Avg'],                         'trigger_flags': ['NV'], 'set_flag': 'NV'},

    # ClimaVUE50 — Dist_km also receives DF when Strikes is R/E/DF.
    {'target': 'Dist_km_Avg',  'sources': ['Strikes_Tot'],                       'trigger_flags': ['R', 'E', 'DF', 'M'], 'set_flag': 'DF'},
    # ClimaVUE50 — lightning distance NV is applied directly in pipeline
    # when Strikes_Tot <= 0 (per notes: Dist valid only if strikes > 0).

    # ClimaVUE50 — max lightning distance depends on Strikes_Tot
    {'target': 'Dist_km_Max',  'sources': ['Strikes_Tot'],                       'trigger_flags': ['R', 'E', 'DF', 'M'], 'set_flag': 'DF'},

    # ClimaVUE50 — time of max lightning distance depends on Strikes_Tot and Dist_km_Avg
    {'target': 'Dist_km_TMx',  'sources': ['Strikes_Tot'],                       'trigger_flags': ['R', 'E', 'DF', 'M'], 'set_flag': 'DF'},
    {'target': 'Dist_km_TMx',  'sources': ['Dist_km_Avg'],                       'trigger_flags': ['R', 'E', 'DF', 'M'], 'set_flag': 'DF'},
    {'target': 'Dist_km_TMx',  'sources': ['Dist_km_Avg'],                       'trigger_flags': ['NV'], 'set_flag': 'NV'},

    # ClimaVUE50 — wind direction standard deviation depends on WindDir
    {'target': 'WindDir_SD1_WVT', 'sources': ['WindDir'],                        'trigger_flags': ['R', 'E', 'DF', 'M'], 'set_flag': 'DF'},
    {'target': 'WindDir_SD1_WVT', 'sources': ['WindDir'],                        'trigger_flags': ['NV'], 'set_flag': 'NV'},

    # ClimaVUE50 — time of max wind gust depends on WS_ms_Avg
    {'target': 'MaxWS_ms_TMx',   'sources': ['WS_ms_Avg'],                      'trigger_flags': ['R', 'E', 'DF', 'M'], 'set_flag': 'DF'},

    # -----------------------------------------------------------------------
    # SR50 Sonic Ranger
    # -----------------------------------------------------------------------
    # DT affects TCDT (DF if DT is R or E)
    {'target': 'TCDT_Avg',     'sources': ['DT_Avg'],                            'trigger_flags': ['R', 'E', 'M'], 'set_flag': 'DF'},
    # Q (quality) affects TCDT — DC if Q is C (uncertain echo)
    {'target': 'TCDT_Avg',     'sources': ['Q_Avg'],                             'trigger_flags': ['C'], 'set_flag': 'DC'},
    # Q affects TCDT — DF if Q is R, E, or DF (bad/no echo)
    {'target': 'TCDT_Avg',     'sources': ['Q_Avg'],                            'trigger_flags': ['R', 'E', 'M'], 'set_flag': 'DF'},
    # AirT affects TCDT (temperature correction) — DC if AirT is DC
    {'target': 'TCDT_Avg',     'sources': ['AirT_C_Avg'],                        'trigger_flags': ['DC'], 'set_flag': 'DC'},
    # AirT affects TCDT — DF if AirT is R, E, or DF
    {'target': 'TCDT_Avg',     'sources': ['AirT_C_Avg'],                        'trigger_flags': ['R', 'E', 'DF', 'M'], 'set_flag': 'DF'},
     # Q depends on DT (Flags_Depend includes DF)
    {'target': 'Q_Avg',        'sources': ['DT_Avg'],                            'trigger_flags': ['R', 'E', 'DF', 'M'], 'set_flag': 'DF'},

    # TCDT affects snow depth (DBTCDT)
    {'target': 'DBTCDT_Avg',   'sources': ['TCDT_Avg'],                          'trigger_flags': ['R', 'E', 'DF', 'M'], 'set_flag': 'DF'},
    # DBTCDT gets DC if TCDT is DC (per Notes: "Flag DC if TCDT == DC")
    {'target': 'DBTCDT_Avg',   'sources': ['TCDT_Avg'],                          'trigger_flags': ['DC'], 'set_flag': 'DC'},

    # -----------------------------------------------------------------------
    # SR50AT Sonic Ranger (with built-in temperature sensor)
    # -----------------------------------------------------------------------
    # SRTmp (built-in temp) affects TCDT — DC if SRTmp is DC
    {'target': 'TCDT_Avg',     'sources': ['SRTmp_C_Avg'],                       'trigger_flags': ['DC'], 'set_flag': 'DC'},
    # SRTmp affects TCDT — DF if SRTmp is R, E, or DF
    {'target': 'TCDT_Avg',     'sources': ['SRTmp_C_Avg'],                       'trigger_flags': ['R', 'E', 'DF', 'M'], 'set_flag': 'DF'},

    # -----------------------------------------------------------------------
    # SN-500 Net Radiometer
    # -----------------------------------------------------------------------
    # SWin/SWout affect SWnet — DF if R or E
    {'target': 'SWnet_Avg',    'sources': ['SWin_Avg', 'SWout_Avg'],             'trigger_flags': ['R', 'E', 'DF', 'M'], 'set_flag': 'DF'},
    # SWnet gets DC if SWin or SWout is C (per Notes: "Flag DC if SWout_Avg OR SWin_Avg == C")
    {'target': 'SWnet_Avg',    'sources': ['SWin_Avg', 'SWout_Avg'],             'trigger_flags': ['C'], 'set_flag': 'DC'},
    # SWnet inherits Z from SWin
    {'target': 'SWnet_Avg',    'sources': ['SWin_Avg'],                          'trigger_flags': ['Z'], 'set_flag': 'Z'},
    # SWout Z is applied directly by nighttime sign checks.
    # LWin/LWout affect LWnet — DF if R or E
    {'target': 'LWnet_Avg',    'sources': ['LWin_Avg', 'LWout_Avg'],             'trigger_flags': ['R', 'E', 'DF', 'M'], 'set_flag': 'DF'},
    # LWnet gets DC if LWin or LWout is C (per Notes: "flag DC if LWin_Avg OR LWout_Avg == C")
    {'target': 'LWnet_Avg',    'sources': ['LWin_Avg', 'LWout_Avg'],             'trigger_flags': ['C'], 'set_flag': 'DC'},

    # SWin/SWout affect albedo — DF if R or E
    {'target': 'SWalbedo_Avg', 'sources': ['SWin_Avg', 'SWout_Avg'],             'trigger_flags': ['R', 'E', 'DF', 'M'], 'set_flag': 'DF'},
    # Albedo gets DC if SWin or SWout is C (per Notes: "Flag DC if SWout_Avg OR SWin_Avg == C")
    {'target': 'SWalbedo_Avg', 'sources': ['SWin_Avg', 'SWout_Avg'],             'trigger_flags': ['C'], 'set_flag': 'DC'},
    # Albedo inherits Z from SWin
    {'target': 'SWalbedo_Avg', 'sources': ['SWin_Avg'],                          'trigger_flags': ['Z'], 'set_flag': 'Z'},
    # DZ is applied programmatically in the pipeline (SWin < 20 W/m²), not via dependency propagation

    # NR_Avg depends on SWnet/LWnet for DF, and uses SWin/SWout/LWin/LWout for DC.
    {'target': 'NR_Avg',       'sources': ['SWnet_Avg', 'LWnet_Avg', 'LWin_Avg', 'LWout_Avg'], 'trigger_flags': ['R', 'E', 'DF', 'M'], 'set_flag': 'DF'},
    # NR gets DC if SWin/SWout/LWin/LWout has C (per RefSensorThresholds Notes)
    {'target': 'NR_Avg',       'sources': ['SWin_Avg', 'SWout_Avg', 'LWin_Avg', 'LWout_Avg'], 'trigger_flags': ['C'], 'set_flag': 'DC'},

    # -----------------------------------------------------------------------
    # SnowVue10 Smart Snow Depth Sensor
    # NOTE: SV_TCDT depends on SV_Tmp (SnowVue internal temp), NOT AirT_C_Avg
    # -----------------------------------------------------------------------
    # SV_Q depends on SV_DT (DF if DT is R, E, or M)
    {'target': 'SV_Q_Avg',      'sources': ['SV_DT_Avg'],                         'trigger_flags': ['R', 'E', 'DF', 'M'], 'set_flag': 'DF'},

    # SV_TCDT depends on SV_Tmp (internal temp, NOT AirT_C_Avg) — DF if R, E, or M
    {'target': 'SV_TCDT_Avg',   'sources': ['SV_Tmp_Avg'],                        'trigger_flags': ['R', 'E', 'DF', 'M'], 'set_flag': 'DF'},
    # SV_TCDT depends on SV_Q — DC if C, DF if R/E/M
    {'target': 'SV_TCDT_Avg',   'sources': ['SV_Q_Avg'],                          'trigger_flags': ['C'], 'set_flag': 'DC'},
    {'target': 'SV_TCDT_Avg',   'sources': ['SV_Q_Avg'],                          'trigger_flags': ['R', 'E', 'M'], 'set_flag': 'DF'},
    # SV_TCDT depends on SV_DT — DF if R/E/M
    {'target': 'SV_TCDT_Avg',   'sources': ['SV_DT_Avg'],                         'trigger_flags': ['R', 'E', 'M'], 'set_flag': 'DF'},

    # SV_DBTCDT (snow depth) depends on SV_TCDT — DF if R/E/DF/M, DC if DC
    {'target': 'SV_DBTCDT_Avg', 'sources': ['SV_TCDT_Avg'],                       'trigger_flags': ['R', 'E', 'DF', 'M'], 'set_flag': 'DF'},
    {'target': 'SV_DBTCDT_Avg', 'sources': ['SV_TCDT_Avg'],                       'trigger_flags': ['DC'], 'set_flag': 'DC'},
    # SV_DBTCDT — T flag when pitch/roll exceeds accuracy range (C flag)
    {'target': 'SV_DBTCDT_Avg', 'sources': ['SV_Pitch', 'SV_Roll'],               'trigger_flags': ['C'], 'set_flag': 'T'},
    # SV_DBTCDT — DF from SV_Alert (E means alert active)
    {'target': 'SV_DBTCDT_Avg', 'sources': ['SV_Alert'],                          'trigger_flags': ['E'], 'set_flag': 'DF'},
    # SV_DBTCDT — DC from SV_Volt, SV_Freq diagnostic C flags
    {'target': 'SV_DBTCDT_Avg', 'sources': ['SV_Volt', 'SV_Freq'],                'trigger_flags': ['C'], 'set_flag': 'DC'},
    # SV_DBTCDT — direct quality dependency on SV_Q (in addition to the SV_TCDT chain)
    {'target': 'SV_DBTCDT_Avg', 'sources': ['SV_Q_Avg'],                          'trigger_flags': ['C'], 'set_flag': 'DC'},
    {'target': 'SV_DBTCDT_Avg', 'sources': ['SV_Q_Avg'],                          'trigger_flags': ['R', 'E', 'DF', 'M'], 'set_flag': 'DF'},
]

# Solar columns that get the nighttime Z-flag check.
# SWnet does NOT get Z from this loop — it inherits Z from SWin via dependency.
SOLAR_COLUMNS = ['SlrFD_W_Avg', 'SWin_Avg', 'SWout_Avg']

# Timestamp-like columns that carry Date/Time values (not numeric).
# These columns skip the threshold loop but still receive dependency/NV/BV/PT/LR flags.
TIMESTAMP_LIKE_COLUMNS = {'MaxWS_ms_TMx', 'Dist_km_TMx'}

# Canonical sensor column names with accepted alias spellings.
COLUMN_ALIASES = {
    'stmp_Avg': 'Stmp_Avg',
    'gtmp_Avg': 'Gtmp_Avg',
    'RHT_Avg': 'RHT_C_Avg',
    'WindDir_SD1': 'WindDir_SD1_WVT',
}

# Threshold-key equivalence for cases where the same variable appears under
# alternate names across files/configs (without renaming raw data columns).
THRESHOLD_KEY_EQUIVALENTS = {
    'MaxWS_ms_Avg': ['MaxWS_ms'],
    'MaxWS_ms': ['MaxWS_ms_Avg'],
    'RHT_C_Avg': ['RHT_Avg'],
    'RHT_Avg': ['RHT_C_Avg'],
    'WindDir_SD1': ['WindDir_SD1_WVT'],
    'WindDir_SD1_WVT': ['WindDir_SD1'],
}

# --- Helper Functions ---

def canonicalize_column_name(name):
    return COLUMN_ALIASES.get(name, name)

def resolve_height_formula_token(value, sensor_height):
    """
    Resolve formula tokens like 'H-50' / 'H+5' against a sensor height.
    Returns the original value when no token is detected.
    """
    if not isinstance(value, str):
        return value
    token = value.strip().upper().replace(" ", "")
    if token == "":
        return value
    if token == "H+5":
        return float(sensor_height) + 5.0
    if token == "H-50":
        return float(sensor_height) - 50.0

    m = re.fullmatch(r"H([+-])(\d+(?:\.\d+)?)", token)
    if not m:
        return value
    op, delta_raw = m.groups()
    delta = float(delta_raw)
    return float(sensor_height) + delta if op == "+" else float(sensor_height) - delta

def threshold_key_variants(col_name):
    """
    Returns equivalent threshold keys to try for a column name.
    """
    variants = []
    for key in [col_name, canonicalize_column_name(col_name)]:
        if key not in variants:
            variants.append(key)

    # Expand with explicit threshold-equivalent names.
    i = 0
    while i < len(variants):
        key = variants[i]
        for alt in THRESHOLD_KEY_EQUIVALENTS.get(key, []):
            if alt not in variants:
                variants.append(alt)
        i += 1
    return variants

def get_threshold_spec_for_column(thresholds, col_name):
    """
    Fetches threshold spec for a column, checking equivalent key variants.
    Returns (spec, matched_key) or (None, None).
    """
    if not isinstance(thresholds, dict):
        return None, None
    for key in threshold_key_variants(col_name):
        if key in thresholds:
            return thresholds[key], key
    return None, None

def normalize_df_column_aliases(df):
    """
    Canonicalize known alias columns in a DataFrame.
    If both alias and canonical columns exist, merge by filling canonical NaNs
    with alias values, then drop the alias column.
    """
    if df is None:
        return df
    for alias, canonical in COLUMN_ALIASES.items():
        if alias not in df.columns:
            continue
        if canonical in df.columns:
            df[canonical] = df[canonical].where(df[canonical].notna(), df[alias])
            df = df.drop(columns=[alias])
        else:
            df = df.rename(columns={alias: canonical})
    return df

def normalize_group_threshold_aliases(groups):
    """
    Canonicalize known alias keys inside instrument-group thresholds.
    """
    if not isinstance(groups, dict):
        return groups
    for _, grp_data in groups.items():
        if not isinstance(grp_data, dict):
            continue
        thresholds = grp_data.get("thresholds", grp_data if isinstance(grp_data, dict) else {})
        if not isinstance(thresholds, dict):
            continue
        for alias, canonical in COLUMN_ALIASES.items():
            if alias not in thresholds:
                continue
            if canonical not in thresholds:
                thresholds[canonical] = thresholds[alias]
            del thresholds[alias]
        if "thresholds" in grp_data:
            grp_data["thresholds"] = thresholds
    return groups

def load_json_file(filepath, default=None):
    if default is None: default = {}
    if os.path.exists(filepath):
        try:
            with open(filepath, 'r') as f:
                return json.load(f)
        except Exception as e:
            st.error(f"Error loading {filepath}: {e}")
            return default
    return default

def save_json_file(filepath, data):
    try:
        with open(filepath, 'w') as f:
            json.dump(data, f, indent=4)
    except Exception as e:
         st.error(f"Error saving {filepath}: {e}")

def load_mapping():
    return load_json_file(MAPPING_FILE, {})

def save_mapping(mapping):
    save_json_file(MAPPING_FILE, mapping)

def load_instrument_groups():
    """
    Loads instrument group configs from GROUPS_FILE.
    If file is empty or missing, seeds it with INITIAL_INSTRUMENT_GROUPS
    (which already contains the full R/C threshold structure).
    """
    groups = load_json_file(GROUPS_FILE, {})
    if not groups:
        # Deep-copy the initial structure so mutations don't affect the constant
        import copy
        groups = copy.deepcopy(INITIAL_INSTRUMENT_GROUPS)
        groups = normalize_group_threshold_aliases(groups)
        save_json_file(GROUPS_FILE, groups)
        return groups
    groups = normalize_group_threshold_aliases(groups)
    save_json_file(GROUPS_FILE, groups)
    return groups


def save_instrument_groups(groups):
    save_json_file(GROUPS_FILE, groups)

def load_station_configs():
    return load_json_file(STATION_CONFIG_FILE, {})

def save_station_configs(configs):
    save_json_file(STATION_CONFIG_FILE, configs)

def parse_toa5_header(file):
    """
    Parses the first line of a TOA5 file (StringIO or UploadedFile)
    or returns default metadata for Excel files.
    Returns a dict with metadata, defaults to '999' if missing.
    """
    meta = {
        'logger_id': '999',
        'logger_script': '999',
        'logger_software': '999'
    }
    
    # Check if it's an Excel file
    if file.name.endswith(('.xlsx', '.xls')):
        st.info(f"Excel file detected: {file.name}. Using default metadata.")
        return meta
    
    try:
        # Read first line without consuming the file permanently
        # For UploadedFile, we read, parse, then seek(0)
        line = file.readline().decode('utf-8').strip()
        parts = [p.strip().strip('"') for p in line.split(',')]
        
        # TOA5 Standard: Format, StationName, Model, Serial, OS, ProgramName, Sig, Table
        if len(parts) >= 6:
            model = parts[2]
            serial = parts[3]
            os_ver = parts[4]
            prog_name = parts[5]
            
            meta['logger_id'] = f"{model}-{serial}"
            meta['logger_software'] = f"{model}-{serial}-{os_ver}"
            meta['logger_script'] = f"{model}-{serial}-{prog_name}" # Version approx
            
    except Exception as e:
        st.warning(f"Could not parse TOA5 header: {e}")
        
    finally:
        file.seek(0)
        
    return meta

def load_csv_preview(file, skip_rows=None):
    """Loads header and first few rows for preview.
    
    Args:
        file: Uploaded file object.
        skip_rows: List of 0-indexed row numbers to skip (default: [0, 2, 3] for TOA5).
    """
    if skip_rows is None:
        skip_rows = [0, 2, 3]  # TOA5 standard: env header, units, processing type
    try:
        df = pd.read_csv(file, skiprows=skip_rows, nrows=5, keep_default_na=False)
        file.seek(0)
        return df
    except Exception as e:
        return None

def parse_field_report(pdf_file):
    """
    Parses a Field Report PDF to extract Date, Time-in, and Time-out.
    Expects formats like:
    Date YYYY-MM-DD
    Time-in HH:MM ...
    Time-out HH:MM ...
    
    Returns a dict with 'field_in', 'field_out' found strings or None.
    """
    if not PdfReader:
        st.error("pypdf not installed. Cannot parse PDF.")
        return {}

    parsed_data = {}
    
    try:
        reader = PdfReader(pdf_file)
        full_text = ""
        
        # Read all pages (usually short reports) to be safe
        for page in reader.pages:
            full_text += page.extract_text() + "\n"
            
        import re
        
        # 1. Extract Date
        # Pattern: Date YYYY-MM-DD
        date_match = re.search(r'Date\s+(\d{4}-\d{2}-\d{2})', full_text)
        report_date = date_match.group(1) if date_match else None
        
        # 2. Extract Time-in
        # Pattern: Time-in HH:MM (ignore rest)
        in_match = re.search(r'Time-in\s+(\d{1,2}:\d{2})', full_text)
        time_in = in_match.group(1) if in_match else None
        
        # 3. Extract Time-out
        # Pattern: Time-out HH:MM
        out_match = re.search(r'Time-out\s+(\d{1,2}:\d{2})', full_text)
        time_out = out_match.group(1) if out_match else None
        
        if report_date and time_in:
            parsed_data['field_in'] = f"{report_date} {time_in}"
            
        if report_date and time_out:
            parsed_data['field_out'] = f"{report_date} {time_out}"
            
    except Exception as e:
        st.warning(f"Error parsing PDF: {e}")
        
    finally:
        pdf_file.seek(0)
        
    return parsed_data

def parse_metadata_log(xlsx_file, raw_filename=None, df_timestamps=None):
    """
    Parses a MetadataLog.xlsx file (EventLog sheet) and extracts three things:

    1. visit_windows  — list of (datetime_in, datetime_out) tuples from 'Site Visit' rows.
                        Only windows whose time_in timestamp actually exists in df_timestamps
                        are returned (prevents flagging periods outside the data file).
                        If time_out < time_in on the same day, the out-time is rolled to
                        the next calendar day (handles the 'Time out is the next day' note).

    2. data_id        — the Data/Visit/Script_ID from the 'Data Download' row whose
                        File_name matches raw_filename (case-insensitive, no extension).
                        Falls back to None if no match is found.

    3. script_id      — the Data/Visit/Script_ID from the most recent 'Script Change' row
                        on or before the download date of the matched Data Download row.
                        Falls back to None if no Script Change rows exist.

    Args:
        xlsx_file    : file-like object for MetadataLog.xlsx
        raw_filename : str, name of the raw data file being processed (used to match Data_ID)
        df_timestamps: pd.Series of datetime values from the raw data file (used to validate
                       that visit windows actually overlap with the file's data range)

    Returns:
        dict with keys 'visit_windows', 'data_id', 'script_id'
    """
    result = {'visit_windows': [], 'data_id': None, 'script_id': None, 'field_time': None}

    # Guard: openpyxl must be available
    if openpyxl is None:
        st.error("openpyxl is not installed. Run: pip install openpyxl")
        return result

    try:
        wb = openpyxl.load_workbook(xlsx_file, data_only=True)

        # The EventLog sheet holds all event rows
        if 'EventLog' not in wb.sheetnames:
            st.warning("MetadataLog.xlsx does not contain an 'EventLog' sheet.")
            return result

        ws = wb['EventLog']

        # Read all rows into a list of dicts keyed by the header row
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Skip completely empty rows
            if all(v is None for v in row):
                continue
            rows.append(dict(zip(headers, row)))

        # Helper: safely parse a time value from the cell (datetime.time or 'NULL' string)
        def _to_time(val):
            """Return a datetime.time or None."""
            if val is None or str(val).strip().upper() == 'NULL':
                return None
            if hasattr(val, 'hour'):  # already a datetime.time
                return val
            # Try parsing string like 'HH:MM'
            try:
                return datetime.strptime(str(val).strip(), '%H:%M').time()
            except ValueError:
                return None

        # Helper: safely parse Date cells that may be true dates, YYYYMMDD ints, or strings
        def _to_date(val):
            """Return a datetime.date or None."""
            if val is None:
                return None

            # datetime/date-like objects
            if hasattr(val, 'date'):
                try:
                    return val.date()
                except Exception:
                    pass

            # Numeric values (e.g., 20250618 or Excel serial day numbers)
            if isinstance(val, (int, np.integer)):
                val_int = int(val)
                as_text = str(val_int)
                if len(as_text) == 8:
                    try:
                        return datetime.strptime(as_text, '%Y%m%d').date()
                    except ValueError:
                        pass
                if 20000 <= val_int <= 60000:
                    try:
                        return (datetime(1899, 12, 30) + timedelta(days=val_int)).date()
                    except Exception:
                        return None
                return None

            if isinstance(val, (float, np.floating)):
                if pd.isna(val):
                    return None
                return _to_date(int(val))

            # Text values
            text = str(val).strip()
            if not text or text.upper() in {'NULL', 'NAN'}:
                return None

            for fmt in ('%Y%m%d', '%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%d/%m/%Y'):
                try:
                    return datetime.strptime(text, fmt).date()
                except ValueError:
                    continue

            parsed = pd.to_datetime(text, errors='coerce')
            if pd.notna(parsed):
                return parsed.date()
            return None

        # Helper: combine a date + time into a datetime, with next-day roll if needed
        def _combine(date_val, time_val, reference_time=None):
            """
            Combine a date (datetime or date) and a time into a datetime.
            If reference_time is provided and time_val < reference_time,
            the date is rolled forward by one day (handles 'next day' time-out).
            """
            if date_val is None or time_val is None:
                return None
            base_date = _to_date(date_val)
            if base_date is None:
                return None
            dt = datetime.combine(base_date, time_val)
            # Auto-roll: if time_out < time_in, the visit ended the next calendar day
            if reference_time is not None and time_val < reference_time:
                dt += timedelta(days=1)
            return dt

        # Determine the data file's timestamp range for validation
        ts_min = df_timestamps.min() if df_timestamps is not None and len(df_timestamps) > 0 else None
        ts_max = df_timestamps.max() if df_timestamps is not None and len(df_timestamps) > 0 else None

        # --- 1. Extract visit windows from 'Site Visit' rows ---
        for row in rows:
            if str(row.get('Event_type', '')).strip() != 'Site Visit':
                continue

            date_val  = _to_date(row.get('Date'))
            time_in   = _to_time(row.get('Time-in'))
            time_out  = _to_time(row.get('Time-out'))

            if date_val is None or time_in is None or time_out is None:
                continue  # skip rows with missing time data

            dt_in  = _combine(date_val, time_in)
            dt_out = _combine(date_val, time_out, reference_time=time_in)

            # Validate: only include this window if it overlaps with the data file's range.
            # If we have no timestamp info, include all windows (conservative).
            if ts_min is not None and ts_max is not None:
                # Window must overlap [ts_min, ts_max]
                if dt_out < ts_min or dt_in > ts_max:
                    continue  # visit is entirely outside this file's date range

            result['visit_windows'].append((dt_in, dt_out))

        # --- 1b. Extract the date embedded in the raw filename (e.g. "20231102" → 2023-11-02) ---
        # This is the most reliable anchor for matching Site Visit rows and pre-filling
        # Field In/Out, because the visit always happens on the same day as the download.
        # We look for an 8-digit YYYYMMDD pattern anywhere in the filename stem.
        import re as _re
        filename_date = None  # datetime.date or None
        raw_stem_for_date = os.path.splitext(os.path.basename(raw_filename or ''))[0]
        _date_match = _re.search(r'(\d{8})', raw_stem_for_date)
        if _date_match:
            try:
                filename_date = datetime.strptime(_date_match.group(1), '%Y%m%d').date()
            except ValueError:
                filename_date = None

        # --- 1c. Find Field In/Out from the Site Visit on the filename date (±2 days) ---
        # We search ALL Site Visit rows directly — NOT filtered through visit_windows —
        # because the overlap filter can exclude visits that start just after ts_max
        # (e.g. visit at 14:33 when the last data point is 14:30).
        for row in rows:
            if str(row.get('Event_type', '')).strip() != 'Site Visit':
                continue
            date_val  = _to_date(row.get('Date'))
            time_in   = _to_time(row.get('Time-in'))
            time_out  = _to_time(row.get('Time-out'))
            if date_val is None or time_in is None or time_out is None:
                continue
            row_date = _to_date(date_val)
            if row_date is None:
                continue

            # Match by filename date (exact or within 2 days), or fall back to ts_max proximity
            matched = False
            if filename_date is not None:
                matched = abs((row_date - filename_date).days) <= 2
            elif ts_max is not None:
                matched = abs((row_date - ts_max.date()).days) <= 2

            if matched:
                dt_in  = _combine(date_val, time_in)
                dt_out = _combine(date_val, time_out, reference_time=time_in)
                result['field_time'] = {'in': dt_in, 'out': dt_out}
                break  # first match wins; user can edit if wrong

        # --- 2. Find Data_ID from 'Data Download' rows (authoritative source) ---
        # Strategy:
        #   a) Prefer the Data Download row whose File_name matches raw_filename.
        #   b) If filename matching fails (legacy naming mismatch), fall back to the
        #      nearest Data Download row date (within +/- 2 days of filename date).
        raw_stem = os.path.splitext(os.path.basename(raw_filename or ''))[0].lower()
        matched_download_date = None

        data_download_rows = [
            r for r in rows
            if str(r.get('Event_type', '')).strip() == 'Data Download'
        ]

        matched_download_row = None

        # (a) Direct filename match
        for row in data_download_rows:
            file_name_cell = str(row.get('File_name', '') or '').strip()
            cell_stem = os.path.splitext(file_name_cell)[0].lower()
            if raw_stem and (raw_stem == cell_stem or raw_stem in cell_stem or cell_stem in raw_stem):
                matched_download_row = row
                break

        # (b) Date-nearest fallback (within +/- 2 days)
        if matched_download_row is None:
            target_date = filename_date if filename_date is not None else (ts_max.date() if ts_max is not None else None)
            if target_date is not None:
                candidates = []
                for row in data_download_rows:
                    date_val = _to_date(row.get('Date'))
                    if date_val is None:
                        continue
                    row_date = _to_date(date_val)
                    if row_date is None:
                        continue
                    day_diff = abs((row_date - target_date).days)
                    if day_diff <= 2:
                        candidates.append((day_diff, row))
                if candidates:
                    candidates.sort(key=lambda x: x[0])
                    matched_download_row = candidates[0][1]

        if matched_download_row is not None:
            result['data_id'] = str(matched_download_row.get('Data/Visit/Script_ID', '') or '').strip() or None
            matched_download_date = _to_date(matched_download_row.get('Date'))

        # --- 3. Find Script_ID from most recent 'Script Change' on or before the matched date ---
        script_rows = [
            r for r in rows
            if str(r.get('Event_type', '')).strip() == 'Script Change'
            and _to_date(r.get('Date')) is not None
        ]

        if script_rows and matched_download_date is not None:
            # Normalise dates for comparison
            download_date = _to_date(matched_download_date)
            # Keep only script changes on or before the matched date
            eligible = [r for r in script_rows if _to_date(r['Date']) <= download_date]
            if eligible:
                # Most recent = largest date
                latest = max(eligible, key=lambda r: _to_date(r['Date']))
                result['script_id'] = str(latest.get('Data/Visit/Script_ID', '') or '').strip() or None

    except Exception as e:
        st.warning(f"Error parsing MetadataLog: {e}")

    return result


def process_file_data(uploaded_file, mapping, metadata, data_id, station_id, skip_rows=None):
    """
    Reads the full CSV or Excel file, applies mapping, adds metadata columns.
    Returns processed DataFrame.

    Args:
        skip_rows: List of 0-indexed row numbers to skip for CSV files.
                   Defaults to [0, 2, 3] (standard TOA5 layout).
    """

    if skip_rows is None:
        skip_rows = [0, 2, 3]  # TOA5: environment row, units row, processing-type row
    try:
        # Check file type
        is_excel = uploaded_file.name.endswith(('.xlsx', '.xls'))
        
        if is_excel:
            # Read Excel file
            # Assume data starts at row 3 (row 1 = headers, row 2 = units)
            df = pd.read_excel(uploaded_file, skiprows=[1],  # Skip units row
                             na_values=['NAN', '"NAN"', '', '-7999', '7999'])
        else:
            # Read CSV/TOA5 file.
            # skip_rows is passed in from the UI; defaults to [0, 2, 3] (TOA5 standard).
            df = pd.read_csv(uploaded_file, skiprows=skip_rows,
                             na_values=['NAN', '"NAN"', '', '-7999', '7999'],
                             keep_default_na=True, skipinitialspace=True, low_memory=False)
        
        # Reset file pointer for next read if needed (though Streamlit handles this usually)
        uploaded_file.seek(0)

        # Rename columns
        # Invert mapping? No, mapping is Source -> Target
        # df.rename(columns=mapping)
        # However, mapping might be incomplete, so only rename known keys
        # 1. Identify Drop Columns
        cols_to_drop = [k for k, v in mapping.items() if v == "REMOVE" and k in df.columns]
        if cols_to_drop:
            df = df.drop(columns=cols_to_drop)
            
        # 2. Rename remaining
        clean_mapping = {k: v for k, v in mapping.items() if k in df.columns and v != k and v != "REMOVE"}
        if clean_mapping:
            df = df.rename(columns=clean_mapping)

        # Canonicalize alias column names (e.g., stmp_Avg -> Stmp_Avg)
        df = normalize_df_column_aliases(df)
            
        # Standardize Timestamp
        if 'TIMESTAMP' in df.columns:
            df['TIMESTAMP'] = pd.to_datetime(df['TIMESTAMP'])

        # Constant station-local UTC offset tag (metadata column, no flag column)
        df['UTC Offset'] = '-07:00'
        
        # Add Metadata Columns
        df['Data_ID'] = str(data_id)
        df['Station_ID'] = station_id
        df['Logger_ID'] = metadata.get('Logger_ID', '999')
        df['Logger_Script'] = metadata.get('Logger_Script', '999')
        df['Logger_Software'] = metadata.get('Logger_Software', '999')
        
        return df

    except Exception as e:
        st.error(f"Error processing file {uploaded_file.name}: {e}")
        return pd.DataFrame()

def write_csv_with_units(df, save_path):
    """
    Writes DataFrame to CSV with a second row containing units.
    Units are looked up from column_mapping.json.
    """
    # Helper to load mapping
    mapping = load_mapping()
    
    with open(save_path, 'w', newline='') as f:
        writer = csv.writer(f)
        # 1. Header
        writer.writerow(df.columns)
        
        # 2. Units
        units_row = []
        for col in df.columns:
            unit_val = "nan" # Default
            
            if col == 'TIMESTAMP': 
                unit_val = 'TS'
            elif col == 'UTC Offset':
                unit_val = 'UTC'
            elif col == 'RECORD': 
                unit_val = 'RN'
            elif col.endswith('_Flag'): 
                unit_val = 'nan'
            else:
                # Lookup in mapping
                if col in mapping:
                    info = mapping[col]
                    if isinstance(info, dict):
                        unit_val = info.get('unit', 'nan')
                        # If blank in JSON, use nan or empty? User sample had 'nan' for flags.
                        # Let's align with observed data "TS,RN,nan,Volts..."
                        if unit_val == "": unit_val = "nan"
            
            units_row.append(unit_val)
        
        writer.writerow(units_row)
        
    # 3. Data
    df.to_csv(save_path, mode='a', header=False, index=False, na_rep='NaN')

@st.cache_data(show_spinner=False)
def load_qc_visualization_data(file_path):
    """
    Loads QA/QC output CSV and prepares an exploded flag table for visualization.
    Returns:
        df: QA/QC data (units row removed if present)
        flag_long: row-wise flag tokens with columns
                   [row_idx, flag_col, flag, variable]
    """
    df = pd.read_csv(file_path, low_memory=False)

    # Remove units row if present, then parse timestamps for time-series charts
    if not df.empty and 'TIMESTAMP' in df.columns:
        if str(df.iloc[0]['TIMESTAMP']) == 'TS':
            df = df.iloc[1:].reset_index(drop=True)
        df['TIMESTAMP'] = pd.to_datetime(df['TIMESTAMP'], errors='coerce')

    flag_cols = [c for c in df.columns if c.endswith('_Flag')]
    empty_flags = pd.DataFrame(columns=['row_idx', 'flag_col', 'flag', 'variable'])
    if not flag_cols:
        return df, empty_flags

    token_series = (
        df[flag_cols]
        .fillna("")
        .astype(str)
        .stack()
        .str.split(",")
        .explode()
        .str.strip()
    )

    token_series = token_series[
        (token_series != "")
        & (~token_series.str.lower().isin(["nan", "none"]))
    ]

    if token_series.empty:
        return df, empty_flags

    flag_long = token_series.reset_index()
    flag_long.columns = ['row_idx', 'flag_col', 'flag']
    flag_long['variable'] = flag_long['flag_col'].str.replace("_Flag", "", regex=False)

    return df, flag_long

def sop_tidy_output_filename(station_code):
    """
    SOP/FileNames convention for historical tidy output:
    [##XX####]_tidy_historical.csv
    """
    safe_station = re.sub(r"[^A-Za-z0-9]", "", str(station_code).strip())
    return f"{safe_station}_tidy_historical.csv"

def _parse_logger_model_serial(logger_id):
    """
    Parse logger ID forms like 'CR350-1379' or 'CR350_1379' -> (CR350, 1379).
    Returns None if not parseable.
    """
    raw = str(logger_id).strip()
    if raw == "" or raw in {"999", "nan", "None"}:
        return None
    m = re.match(r"^([A-Za-z0-9]+)[-_]([A-Za-z0-9]+)$", raw)
    if not m:
        return None
    return m.group(1), m.group(2)

def build_sop_tidy_save_name(df_qc, station_code, output_dir, selected_input_name=None):
    """
    Build SOP-compliant tidy output name.

    Preferred (contemporary): [STATION]_tidy_[MODEL]_[SERIAL]_[YYYYMMDD].csv
    Fallback (historical):    [STATION]_tidy_historical.csv
    """
    safe_station = re.sub(r"[^A-Za-z0-9]", "", str(station_code).strip())
    fallback_name = sop_tidy_output_filename(safe_station)

    if df_qc is None or not isinstance(df_qc, pd.DataFrame) or 'Logger_ID' not in df_qc.columns:
        return fallback_name

    # Determine a single logger model/serial pair.
    logger_pairs = {
        parsed for parsed in (
            _parse_logger_model_serial(v)
            for v in df_qc['Logger_ID'].dropna().astype(str).tolist()
        )
        if parsed is not None
    }
    if len(logger_pairs) != 1:
        return fallback_name
    model, serial = next(iter(logger_pairs))

    # Determine retrieval date from file names available in the workspace.
    date_candidates = set()
    names_to_scan = []
    if selected_input_name:
        names_to_scan.append(os.path.basename(str(selected_input_name)))
    try:
        names_to_scan.extend(os.listdir(output_dir))
    except Exception:
        pass

    p_specific = re.compile(
        rf"^{re.escape(safe_station)}_(?:raw|tidy)_{re.escape(model)}_{re.escape(serial)}_(\d{{8}})\.(?:csv|dat)$",
        re.IGNORECASE
    )
    for name in names_to_scan:
        m = p_specific.match(os.path.basename(str(name)))
        if m:
            date_candidates.add(m.group(1))

    # If we still can't find one, allow station-level dates from raw/tidy names.
    if not date_candidates:
        p_station = re.compile(
            rf"^{re.escape(safe_station)}_(?:raw|tidy)_[A-Za-z0-9]+_[A-Za-z0-9]+_(\d{{8}})\.(?:csv|dat)$",
            re.IGNORECASE
        )
        for name in names_to_scan:
            m = p_station.match(os.path.basename(str(name)))
            if m:
                date_candidates.add(m.group(1))

    if not date_candidates:
        return fallback_name

    retrieval_date = max(date_candidates)
    return f"{safe_station}_tidy_{model}_{serial}_{retrieval_date}.csv"

def is_qaqc_output_file(filename):
    """
    Accept both legacy and SOP-style QA/QC output filenames.
    """
    name = str(filename)
    # Legacy app output
    if name.endswith("_QC.csv"):
        return True
    # SOP/FileNames tidy outputs (historical or contemporary) saved as CSV
    if re.fullmatch(r"[A-Za-z0-9]+_tidy_historical\.csv", name):
        return True
    if re.fullmatch(r"[A-Za-z0-9]+_tidy_[A-Za-z0-9]+_[A-Za-z0-9]+_\d{8}\.csv", name):
        return True
    return False

def build_qc_viz_report_xlsx(summary_df, flag_counts, variable_counts, matrix, total_over_time=None, by_flag_over_time=None):
    """
    Builds an Excel report for the QA/QC visualization tab.
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        flag_counts.to_excel(writer, sheet_name="Flag Counts", index=False)
        variable_counts.to_excel(writer, sheet_name="Variable Counts", index=False)
        matrix.reset_index().to_excel(writer, sheet_name="Variable x Flag", index=False)

        if total_over_time is not None and not total_over_time.empty:
            total_over_time.reset_index().to_excel(writer, sheet_name="Time Trend Total", index=False)

        if by_flag_over_time is not None and not by_flag_over_time.empty:
            by_flag_over_time.reset_index().to_excel(writer, sheet_name="Time Trend By Flag", index=False)

    output.seek(0)
    return output.getvalue()

FLAG_DISPLAY_NAMES = {
    'C': 'Caution',
    'P': 'Pass',
    'DF': 'Dependency Failure',
    'DC': 'Dependency Caution',
    'DZ': 'Divide by Zero',
    'E': 'Error',
    'ERR': 'Error',
    'M': 'Missing',
    'V': 'Visit',
    'Z': 'Night time Value',
    'T': 'Tilt',
    'SF': 'Snow Free',
    'R': 'Range',
    'NV': 'No Value',
    'BV': 'Battery Voltage',
    'PT': 'Panel Temperature',
    'LR': 'Logger Restart',
}


def format_flag_label(flag_code):
    code = str(flag_code).strip()
    full_name = FLAG_DISPLAY_NAMES.get(code)
    return f"{code} - {full_name}" if full_name else code


def with_display_flag_columns(by_flag_over_time):
    if by_flag_over_time is None or by_flag_over_time.empty:
        return by_flag_over_time
    return by_flag_over_time.rename(columns={c: format_flag_label(c) for c in by_flag_over_time.columns})


def compute_flag_trend_tables(filtered_flags, df_viz, freq_code):
    """
    Builds total and per-flag time trend tables for selected flags.
    """
    trend = filtered_flags[['row_idx', 'flag']].join(
        df_viz[['TIMESTAMP']],
        on='row_idx',
        how='left'
    )
    trend = trend.dropna(subset=['TIMESTAMP'])

    if trend.empty:
        return pd.DataFrame(), pd.DataFrame()

    total_over_time = (
        trend
        .groupby(pd.Grouper(key='TIMESTAMP', freq=freq_code))
        .size()
        .rename("Count")
        .to_frame()
    )
    by_flag_over_time = (
        trend
        .groupby([pd.Grouper(key='TIMESTAMP', freq=freq_code), 'flag'])
        .size()
        .unstack(fill_value=0)
    )
    return total_over_time, by_flag_over_time

def compute_daily_variable_flag_percent_table(filtered_flags, df_viz):
    """
    Builds daily flag percentages at the variable level.

    Output columns:
      TIMESTAMP (day), variable, flag, Flag_Count, Variable_Daily_Count,
      Pct_of_Variable_Daily_Count
    """
    if 'TIMESTAMP' not in df_viz.columns:
        return pd.DataFrame()

    trend = filtered_flags[['row_idx', 'variable', 'flag']].join(
        df_viz[['TIMESTAMP']],
        on='row_idx',
        how='left'
    )
    trend = trend.dropna(subset=['TIMESTAMP'])
    if trend.empty:
        return pd.DataFrame()

    trend['TIMESTAMP'] = pd.to_datetime(trend['TIMESTAMP'], errors='coerce')
    trend = trend.dropna(subset=['TIMESTAMP'])
    if trend.empty:
        return pd.DataFrame()

    numerator = (
        trend
        .groupby([pd.Grouper(key='TIMESTAMP', freq='D'), 'variable', 'flag'])
        .size()
        .rename('Flag_Count')
        .reset_index()
    )

    # Denominator per day+variable:
    # count rows where that variable has a non-null value on that day.
    variables = sorted(numerator['variable'].unique().tolist())
    value_cols = [v for v in variables if v in df_viz.columns]
    valid_ts_mask = df_viz['TIMESTAMP'].notna()

    denominator = pd.DataFrame(columns=['TIMESTAMP', 'variable', 'Variable_Daily_Count'])
    if value_cols and valid_ts_mask.any():
        day_key = pd.to_datetime(df_viz.loc[valid_ts_mask, 'TIMESTAMP']).dt.floor('D')
        denom_matrix = df_viz.loc[valid_ts_mask, value_cols].notna().groupby(day_key).sum()
        denominator = (
            denom_matrix
            .stack()
            .rename('Variable_Daily_Count')
            .reset_index()
            .rename(columns={'level_1': 'variable'})
        )
        denominator['Variable_Daily_Count'] = denominator['Variable_Daily_Count'].astype(int)

    # Fallback denominator for variables not present as data columns:
    # use total timestamped rows for the day.
    day_rows = (
        df_viz.loc[valid_ts_mask, ['TIMESTAMP']]
        .groupby(pd.Grouper(key='TIMESTAMP', freq='D'))
        .size()
        .rename('Daily_Row_Count')
        .reset_index()
    )
    base_pairs = numerator[['TIMESTAMP', 'variable']].drop_duplicates()
    denominator_full = (
        base_pairs
        .merge(denominator, on=['TIMESTAMP', 'variable'], how='left')
        .merge(day_rows, on='TIMESTAMP', how='left')
    )
    denominator_full['Variable_Daily_Count'] = (
        denominator_full['Variable_Daily_Count']
        .fillna(denominator_full['Daily_Row_Count'])
        .fillna(0)
        .astype(int)
    )

    out = numerator.merge(
        denominator_full[['TIMESTAMP', 'variable', 'Variable_Daily_Count']],
        on=['TIMESTAMP', 'variable'],
        how='left'
    )
    out['Pct_of_Variable_Daily_Count'] = np.where(
        out['Variable_Daily_Count'] > 0,
        (out['Flag_Count'] / out['Variable_Daily_Count']) * 100.0,
        0.0
    ).round(2)

    return out.sort_values(['TIMESTAMP', 'variable', 'flag']).reset_index(drop=True)

def build_trend_png(total_over_time, by_flag_over_time, title):
    """
    Renders trend tables as a PNG image (two panels).
    """
    fig, axes = plt.subplots(2, 1, figsize=(12, 8), constrained_layout=True)

    total_over_time.plot(ax=axes[0], legend=False, linewidth=2)
    axes[0].set_title(f"{title} - Total")
    axes[0].set_xlabel("Time")
    axes[0].set_ylabel("Count")
    axes[0].grid(alpha=0.3)

    if not by_flag_over_time.empty:
        by_flag_plot = with_display_flag_columns(by_flag_over_time)
        by_flag_plot.plot(ax=axes[1], linewidth=1.6)
        axes[1].legend(title="Flag (Code - Meaning)", loc="upper right", fontsize=8)
    axes[1].set_title(f"{title} - By Flag")
    axes[1].set_xlabel("Time")
    axes[1].set_ylabel("Count")
    axes[1].grid(alpha=0.3)

    png_buffer = io.BytesIO()
    fig.savefig(png_buffer, format="png", dpi=180)
    plt.close(fig)
    png_buffer.seek(0)
    return png_buffer.getvalue()

def build_daily_variable_flag_pct_png(pct_pivot, title):
    """
    Renders a daily variable-level percentage trend chart as a PNG.
    Expects columns to represent flags for a single variable.
    """
    fig, ax = plt.subplots(figsize=(13, 6), constrained_layout=True)
    pct_pivot.plot(ax=ax, linewidth=1.5)
    ax.set_title(title)
    ax.set_xlabel("Date")
    ax.set_ylabel("% of Variable Records (Daily)")
    ax.set_ylim(bottom=0)
    ax.grid(alpha=0.3)
    if pct_pivot.shape[1] <= 20:
        ax.legend(title="Flag", loc="upper right", fontsize=8)
    else:
        ax.legend([], [], frameon=False)

    png_buffer = io.BytesIO()
    fig.savefig(png_buffer, format="png", dpi=180)
    plt.close(fig)
    png_buffer.seek(0)
    return png_buffer.getvalue()

def build_15min_variable_flag_overlay_png(df_viz, filtered_flags, variable, title, start_ts=None, end_ts=None):
    """
    Render a 15-minute variable time series with flagged points overlaid by flag code.
    Returns PNG bytes, or None when no plottable rows exist in the selected range.
    """
    if (
        df_viz is None
        or filtered_flags is None
        or 'TIMESTAMP' not in df_viz.columns
        or variable not in df_viz.columns
    ):
        return None

    ts = pd.to_datetime(df_viz['TIMESTAMP'], errors='coerce')
    vals = pd.to_numeric(df_viz[variable], errors='coerce')
    base = pd.DataFrame({'TIMESTAMP': ts, 'Value': vals}, index=df_viz.index)
    base = base.dropna(subset=['TIMESTAMP', 'Value'])

    if start_ts is not None:
        base = base[base['TIMESTAMP'] >= pd.Timestamp(start_ts)]
    if end_ts is not None:
        base = base[base['TIMESTAMP'] <= pd.Timestamp(end_ts)]
    if base.empty:
        return None

    points = filtered_flags[filtered_flags['variable'] == variable].copy()
    points = points[points['row_idx'].isin(base.index)]
    if not points.empty:
        base_with_idx = base.reset_index().rename(columns={'index': 'row_idx'})
        points = points.merge(
            base_with_idx[['row_idx', 'TIMESTAMP', 'Value']],
            on='row_idx',
            how='left'
        ).dropna(subset=['TIMESTAMP', 'Value'])

    fig, ax = plt.subplots(figsize=(13, 5.5), constrained_layout=True)
    ax.plot(base['TIMESTAMP'], base['Value'], color="#5a5a5a", linewidth=1.1, label=variable)

    if not points.empty:
        cmap = plt.cm.get_cmap('tab10')
        for i, flag_code in enumerate(sorted(points['flag'].dropna().astype(str).unique().tolist())):
            pf = points[points['flag'] == flag_code]
            if pf.empty:
                continue
            ax.scatter(
                pf['TIMESTAMP'],
                pf['Value'],
                s=18,
                alpha=0.85,
                color=cmap(i % 10),
                label=format_flag_label(flag_code)
            )
        ax.legend(loc="upper right", fontsize=8, title="Flag")

    ax.set_title(title)
    ax.set_xlabel("Time (15-min)")
    ax.set_ylabel(variable)
    ax.grid(alpha=0.3)

    png_buffer = io.BytesIO()
    fig.savefig(png_buffer, format="png", dpi=180)
    plt.close(fig)
    png_buffer.seek(0)
    return png_buffer.getvalue()

# --- UI Components ---

def mapping_editor_ui():
    """
    UI for editing the column mapping JSON using a table editor.
    """
    st.write("Edit the column mapping configuration below.")
    st.info("💡 **Tips:**\n- **Aliases**: Separate multiple values with a comma (e.g. `battv, BattV`).\n- **Add Row**: Click the `+` at the bottom to add a new variable.\n- **Remove Row**: Select rows and press `Delete`.")
    
    current_mapping = load_mapping()
    
    # 1. FLATTEN: Convert JSON Dict -> List of Dicts for DataFrame
    # Structure: Variable Name | Unit | Aliases (str) | Unit Aliases (str)
    table_data = []
    
    for var_name, info in current_mapping.items():
        # Handle cases where info might be malformed (though unlikely with our control)
        if not isinstance(info, dict): continue
        
        aliases_list = info.get('aliases', [])
        unit = info.get('unit', '')
        unit_aliases_list = info.get('unit_aliases', [])
        
        # Convert lists to comma-separated strings for easy editing
        aliases_str = ", ".join([str(a) for a in aliases_list])
        unit_aliases_str = ", ".join([str(ua) for ua in unit_aliases_list])
        
        table_data.append({
            "Variable Name": var_name,
            "Unit": unit,
            "Aliases": aliases_str,
            "Unit Aliases": unit_aliases_str
        })
        
    # Create DataFrame
    df = pd.DataFrame(table_data)
    
    # 2. EDIT: Show Data Editor
    # num_rows="dynamic" allows adding/deleting rows
    edited_df = st.data_editor(
        df, 
        key="column_mapping_editor",
        num_rows="dynamic", 
        use_container_width=True,
        hide_index=True,
        column_config={
            "Variable Name": st.column_config.TextColumn("Variable Name", required=True, help="Canonical name of the variable (e.g. BattV_Avg)"),
            "Unit": st.column_config.TextColumn("Unit", help="Canonical unit (e.g. Volts)"),
            "Aliases": st.column_config.TextColumn("Aliases", help="Comma-separated list of alternative names found in raw files"),
            "Unit Aliases": st.column_config.TextColumn("Unit Aliases", help="Comma-separated list of alternative units found in raw files"),
        }
    )
    
    # 3. SAVE: Parse back to JSON
    if st.button("Save Changes"):
        try:
            new_mapping = {}
            for index, row in edited_df.iterrows():
                var_name = str(row.get("Variable Name", "")).strip()
                if not var_name: 
                    continue # Skip empty variable names
                
                # Parse Aliases
                aliases_raw = str(row.get("Aliases", ""))
                aliases = [a.strip() for a in aliases_raw.split(',') if a.strip()]
                
                # Parse Unit Aliases
                unit_aliases_raw = str(row.get("Unit Aliases", ""))
                unit_aliases = [ua.strip() for ua in unit_aliases_raw.split(',') if ua.strip()]
                
                unit = str(row.get("Unit", "")).strip()
                
                new_mapping[var_name] = {
                    "aliases": aliases,
                    "unit": unit,
                    "unit_aliases": unit_aliases
                }
                
            # Check if empty (safety check)
            if not new_mapping:
                st.warning("Mapping is empty. Are you sure?")
                if not st.button("Confirm Empty Save"):
                    return
            
            save_mapping(new_mapping)
            st.success("Configuration saved successfully!")
            st.rerun()
            
        except Exception as e:
            st.error(f"Error saving configuration: {e}")

def mapping_editor_dialog():
    # Keep this editor outside st.dialog to avoid FragmentStorageKeyError when
    # adding/removing rows in st.data_editor(num_rows="dynamic").
    st.sidebar.markdown("---")
    st.sidebar.subheader("Edit Column Mapping")
    with st.sidebar.expander("Show Editor", expanded=False):
        mapping_editor_ui()

# --- Main App ---

def main():
    st.title("NHG Weather Data Pipeline")

    # --- Sidebar ---
    st.sidebar.header("Global Settings")
    
    # Stable inline editor (avoids fragment lifecycle issues in modal dialog).
    mapping_editor_dialog()

    st.sidebar.divider()

    station_name = st.sidebar.text_input("Station Name", value="Station_Name")
    
    # Latitude and Longitude inputs (only show when station name is not default)
    latitude = 53.7217  # Default values
    longitude = -125.6417
    
    if station_name != "Station_Name":
        # Load station configs to get saved lat/lon
        station_configs = load_station_configs()
        
        # Initialize station if it doesn't exist
        if station_name not in station_configs:
            station_configs[station_name] = {
                "latitude": 53.7217,
                "longitude": -125.6417,
                "deployments": []
            }
            save_station_configs(station_configs)
        
        # Get saved values
        station_data = station_configs[station_name]
        saved_lat = station_data.get("latitude", 53.7217)
        saved_lon = station_data.get("longitude", -125.6417)
        
        st.sidebar.caption("📍 Station Coordinates (for solar calculations)")
        col1, col2 = st.sidebar.columns(2)
        with col1:
            latitude = st.number_input(
                "Latitude", 
                min_value=-90.0, 
                max_value=90.0, 
                value=float(saved_lat),
                step=0.0001,
                format="%.4f",
                key="station_latitude"
            )
        with col2:
            longitude = st.number_input(
                "Longitude", 
                min_value=-180.0, 
                max_value=180.0, 
                value=float(saved_lon),
                step=0.0001,
                format="%.4f",
                key="station_longitude"
            )
        
        # Auto-save if values changed
        if latitude != saved_lat or longitude != saved_lon:
            station_configs[station_name]["latitude"] = latitude
            station_configs[station_name]["longitude"] = longitude
            save_station_configs(station_configs)
    
    # ------------------------------------------------------------------
    # MetadataLog — station-level upload (shared across all uploaded files)
    # Upload once per station; the app parses it and auto-fills Data ID,
    # Logger Script, and Field Visit windows for every file in the session.
    # ------------------------------------------------------------------
    st.sidebar.divider()
    st.sidebar.caption("MetadataLog (Optional)")
    sidebar_metalog = st.sidebar.file_uploader(
        "Upload MetadataLog.xlsx",
        type=["xlsx"],
        key="sidebar_metalog",
        help="Upload once per station. Auto-fills Data ID, Logger Script, and Field Visit windows for all uploaded files."
    )

    if sidebar_metalog:
        # Cache the raw bytes so re-runs don't require re-uploading.
        # If the user swaps to a different file, the name changes and we re-cache.
        if (
            "metalog_raw" not in st.session_state
            or st.session_state.get("metalog_name") != sidebar_metalog.name
        ):
            st.session_state["metalog_raw"]  = sidebar_metalog.read()
            st.session_state["metalog_name"] = sidebar_metalog.name
            sidebar_metalog.seek(0)
        st.sidebar.success(f"Loaded: **{sidebar_metalog.name}**")
    else:
        # User removed the file — clear the cache
        st.session_state.pop("metalog_raw",  None)
        st.session_state.pop("metalog_name", None)

    st.sidebar.divider()

    output_dir = st.sidebar.text_input("Output Directory", value="data")

    # Sensor height is now configured per instrument group
    
    if not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    # --- Tabs ---
    tab1, tab2, tab3, tab4 = st.tabs([
        "1. Ingestion & Concatenation",
        "2. QA/QC Processing",
        "3. QA/QC Visualization",
        "4. Save Trend Graphs"
    ])

    # --- Tab 1: Ingestion ---
    with tab1:
        st.header("1. Ingestion & Standardization")
        
        uploaded_files = st.file_uploader("Upload Raw Logger Files (CSV/TOA5/Excel)", accept_multiple_files=True, type=["csv", "dat", "xlsx", "xls"])
        
        processed_file_configs = [] # specific configs for each file

        if uploaded_files:
            st.divider()
            
            # Load current mapping to help with defaults
            mapping_ref = load_mapping()
            
            # --- Per File Configuration ---
            for i, file in enumerate(uploaded_files):
                with st.expander(f"File {i+1}: {file.name}", expanded=True):
                    col1, col2 = st.columns(2)
                    
                    # Parse TOA5 header for logger metadata (model, serial, OS, script)
                    meta = parse_toa5_header(file)

                    # --- Advanced: Row Skip Settings ---
                    # TOA5 files have 4 rows before data:
                    #   Row 0: environment/file-info header
                    #   Row 1: column names (pandas header — NOT skipped)
                    #   Row 2: units
                    #   Row 3: processing type (Smp, Avg, …)
                    # Use the toggle below to override if your file has extra preamble rows.
                    # Note: st.expander cannot be nested inside another expander in Streamlit,
                    # so we use a toggle to show/hide this setting instead.
                    show_skip_settings = st.toggle(
                        "⚙️ Advanced: Row Skip Settings",
                        value=False,
                        key=f"show_skip_{i}"
                    )
                    if show_skip_settings:
                        st.caption(
                            "Rows to skip before the column-name row. "
                            "Standard TOA5 files need rows 0, 2 and 3 skipped. "
                            "Add extra row numbers if your file has additional headers."
                        )
                        skip_rows = st.multiselect(
                            "Skip rows (0-indexed)",
                            options=list(range(10)),
                            default=[0, 2, 3],
                            key=f"skip_rows_{i}",
                            help="Row 0 = first line of the file. Standard TOA5 = [0, 2, 3]."
                        )
                    else:
                        skip_rows = [0, 2, 3]  # standard TOA5 default

                    # ------------------------------------------------------------------
                    # MetadataLog Auto-fill
                    # The MetadataLog is uploaded once in the sidebar (station-level).
                    # Here we parse it per-file using that file's timestamp range to:
                    #   - Match the Data ID from the 'Data Download' row for this file
                    #   - Find the most recent 'Script Change' row
                    #   - Find the first 'Site Visit' within 2 days of this file's range
                    #     and pre-fill Field In / Field Out (both remain editable)
                    # ------------------------------------------------------------------

                    # Defaults before any auto-fill
                    auto_data_id       = ""
                    auto_script_id     = meta['logger_script']  # fall back to TOA5 header
                    auto_visit_windows = []  # list of (datetime_in, datetime_out)
                    auto_field_in      = ""  # pre-fill for Field In text input
                    auto_field_out     = ""  # pre-fill for Field Out text input

                    if "metalog_raw" in st.session_state:
                        # Quick read of just the TIMESTAMP column to validate visit windows
                        # against this specific file's date range.
                        try:
                            file.seek(0)
                            _ts_df = pd.read_csv(
                                file,
                                skiprows=skip_rows,   # honour the user's row-skip selection
                                usecols=['TIMESTAMP'],
                                na_values=['NAN', '"NAN"', ''],
                                keep_default_na=True,
                                low_memory=False
                            )
                            _ts_series = pd.to_datetime(_ts_df['TIMESTAMP'], errors='coerce').dropna()
                            file.seek(0)  # reset for the full read later
                        except Exception:
                            _ts_series = pd.Series([], dtype='datetime64[ns]')
                            file.seek(0)

                        # Parse the shared MetadataLog bytes for this file's context
                        parsed_meta = parse_metadata_log(
                            io.BytesIO(st.session_state["metalog_raw"]),
                            raw_filename=file.name,
                            df_timestamps=_ts_series
                        )

                        # Apply auto-filled values
                        if parsed_meta['data_id']:
                            auto_data_id = parsed_meta['data_id']
                            st.success(f"✅ Data ID auto-filled: **{auto_data_id}**")
                        if parsed_meta['script_id']:
                            auto_script_id = parsed_meta['script_id']
                            st.success(f"✅ Logger Script auto-filled: **{auto_script_id}**")

                        # Also harvest ALL Script Change IDs from the MetadataLog so the
                        # dropdown shows the full history, not just the best-match entry.
                        # This ensures scripts whose change date predates the file are still
                        # available in the dropdown.
                        try:
                            import io as _io
                            import openpyxl as _opx
                            _wbl = _opx.load_workbook(
                                _io.BytesIO(st.session_state["metalog_raw"]), data_only=True
                            )
                            if 'EventLog' in _wbl.sheetnames:
                                _ws = _wbl['EventLog']
                                _hdrs = [c.value for c in _ws[1]]
                                for _row in _ws.iter_rows(min_row=2, values_only=True):
                                    _rd = dict(zip(_hdrs, _row))
                                    if str(_rd.get('Event_type', '')).strip() == 'Script Change':
                                        _sid = str(_rd.get('Data/Visit/Script_ID', '') or '').strip()
                                        if _sid and _sid.upper() not in ('NULL', '999', ''):
                                            # Store in session state so it persists across files
                                            st.session_state.setdefault(
                                                "known_logger_scripts", set()
                                            ).add(_sid)
                        except Exception:
                            pass  # silently skip if MetadataLog can't be re-read here

                        auto_visit_windows = parsed_meta['visit_windows']
                        if auto_visit_windows:
                            st.success(
                                f"✅ {len(auto_visit_windows)} field visit window(s) found "
                                f"overlapping this file's date range."
                            )

                        # Pre-fill Field In/Out from the first Site Visit within 2 days
                        if parsed_meta.get('field_time'):
                            ft = parsed_meta['field_time']
                            auto_field_in  = ft['in'].strftime('%Y-%m-%d %H:%M')
                            auto_field_out = ft['out'].strftime('%Y-%m-%d %H:%M')
                            st.info(
                                f"ℹ️ Field times pre-filled from MetadataLog visit on "
                                f"{ft['in'].strftime('%Y-%m-%d')}. Edit below if needed."
                            )

                    # ------------------------------------------------------------------
                    # Editable metadata fields (auto-filled above, but user can override)
                    # ------------------------------------------------------------------
                    with col1:
                        # Data ID: auto-filled from MetadataLog, editable override
                        did_val = st.text_input(
                            f"Data ID ({file.name})",
                            value=auto_data_id,
                            key=f"did_{i}",
                            help="Auto-filled from MetadataLog 'Data Download' row. Edit if needed."
                        )
                        data_id = did_val.strip() if did_val.strip() else "999"

                        logger_id = st.text_input(f"Logger ID", value=meta['logger_id'], key=f"lid_{i}")

                    with col2:
                        # Logger Script: dropdown with all known scripts seen so far,
                        # plus a "Custom…" option that reveals a free-text field.
                        # The list is seeded from: TOA5-parsed value + MetadataLog auto-fill.
                        # We store known scripts in session state so they accumulate across files.
                        if "known_logger_scripts" not in st.session_state:
                            st.session_state["known_logger_scripts"] = set()

                        # Add both the TOA5-derived and MetadataLog-derived IDs to the pool
                        for _s in [meta['logger_script'], auto_script_id]:
                            if _s and _s != '999':
                                st.session_state["known_logger_scripts"].add(_s)

                        _CUSTOM_LABEL = "✏️ Custom (type below)…"
                        _script_dropdown_options = (
                            sorted(st.session_state["known_logger_scripts"]) + [_CUSTOM_LABEL]
                        )

                        # Pre-select the auto-filled value if it's in the list
                        _default_idx = (
                            _script_dropdown_options.index(auto_script_id)
                            if auto_script_id in _script_dropdown_options
                            else 0
                        )

                        _selected_script = st.selectbox(
                            "Logger Script",
                            options=_script_dropdown_options,
                            index=_default_idx,
                            key=f"lsc_sel_{i}",
                            help="Auto-filled from MetadataLog or TOA5 header. Choose from the list or select 'Custom' to type a value."
                        )

                        if _selected_script == _CUSTOM_LABEL:
                            # Reveal a free-text field when 'Custom' is chosen
                            logger_script = st.text_input(
                                "Custom Logger Script ID",
                                value=auto_script_id,
                                key=f"lsc_custom_{i}",
                                help="Type any script identifier you need."
                            )
                        else:
                            logger_script = _selected_script
                        logger_soft = st.text_input(f"Logger Software", value=meta['logger_software'], key=f"lsw_{i}")

                    # Caution Flag Option (per-file and column-selectable)
                    add_caution = st.checkbox(
                        f"Add Caution Flag (C) for this dataset",
                        key=f"caution_{i}",
                        help="When enabled, choose which mapped parameters should receive C."
                    )
                    selected_caution_columns = []

                    # ------------------------------------------------------------------
                    # Field Visit Windows (V flag)
                    # Auto-detected windows from MetadataLog are shown as read-only.
                    # Field In / Field Out are pre-filled from the closest matching visit
                    # (within 2 days of this file's date range) and are always editable.
                    # ------------------------------------------------------------------
                    st.caption("🏕️ Field Visit Windows (V flag)")

                    # Show all auto-detected windows as read-only reference
                    if auto_visit_windows:
                        st.write("**All visits overlapping this file's date range:**")
                        for vw_in, vw_out in auto_visit_windows:
                            st.write(f"  • {vw_in.strftime('%Y-%m-%d %H:%M')} → {vw_out.strftime('%Y-%m-%d %H:%M')}")

                    # Field In / Field Out — always shown, pre-filled if a match was found
                    fv_col1, fv_col2 = st.columns(2)
                    with fv_col1:
                        fv_in = st.text_input(
                            "Field In (YYYY-MM-DD HH:MM)",
                            value=auto_field_in,
                            key=f"fi_{i}",
                            help="Pre-filled from MetadataLog Site Visit within 2 days of this file. Edit if needed."
                        )
                    with fv_col2:
                        fv_out = st.text_input(
                            "Field Out (YYYY-MM-DD HH:MM)",
                            value=auto_field_out,
                            key=f"fo_{i}",
                            help="Pre-filled from MetadataLog Site Visit within 2 days of this file. Edit if needed."
                        )


                    
                    # Preview & Mapping
                    # Use the user-selected skip_rows so the preview matches the actual read
                    df_preview = load_csv_preview(file, skip_rows=skip_rows)
                    
                    final_mapping = {}
                    
                    if df_preview is not None:
                        st.dataframe(df_preview.head(), use_container_width=True)
                        
                        columns = df_preview.columns.tolist()
                        
                        # Auto-Map Logic
                        current_mapping = {}
                        for col in columns:
                            target = canonicalize_column_name(col)
                            # Try to find match in JSON
                            for std_name, info in mapping_ref.items():
                                # Handle new dict format (with units) or old list format
                                if isinstance(info, dict):
                                    aliases = info.get('aliases', [])
                                else:
                                    aliases = info # Fallback for old format

                                if col == std_name:
                                    target = std_name
                                    break
                                if col in aliases:
                                    target = std_name
                                    break
                            current_mapping[col] = target

                        st.caption("Column Mapping (Uncheck 'Include' to remove column)")
                        map_df = pd.DataFrame({
                            "Include": [True] * len(columns),
                            "Source": columns,
                            "Target": [current_mapping[c] for c in columns]
                        })
                        
                        # Editable dataframe
                        edited_map = st.data_editor(
                            map_df, 
                            key=f"map_editor_{i}",
                            num_rows="fixed",
                            column_config={
                                "Include": st.column_config.CheckboxColumn(
                                    label="Keep?",
                                    default=True
                                ),
                                "Source": st.column_config.TextColumn(disabled=True),
                                "Target": st.column_config.SelectboxColumn(
                                    label="Target Column",
                                    options=list(set(list(mapping_ref.keys()) + columns)), # Unique options
                                    required=True
                                )
                            },
                            use_container_width=True
                        )
                        
                        # Create dictionary from editor
                        # If Include is False, map to "REMOVE" which is handled in process_file_data
                        for index, row in edited_map.iterrows():
                            if row['Include']:
                                final_mapping[row['Source']] = row['Target']
                            else:
                                final_mapping[row['Source']] = "REMOVE"
                    
                    # Optional per-file caution-column selection (after mapping is known)
                    if add_caution:
                        mapped_targets = sorted({
                            canonicalize_column_name(str(v).strip())
                            for v in final_mapping.values()
                            if v not in [None, "", "REMOVE"]
                            and str(v).strip() not in [
                                'TIMESTAMP', 'UTC Offset', 'RECORD',
                                'Data_ID', 'Station_ID', 'Logger_ID',
                                'Logger_Script', 'Logger_Software'
                            ]
                            and not str(v).strip().endswith('_Flag')
                        })
                        caution_options = mapped_targets
                        if caution_options:
                            selected_caution_columns = st.multiselect(
                                "Select parameters to receive C for this dataset",
                                options=caution_options,
                                default=caution_options,
                                key=f"caution_cols_{i}",
                                help="Only selected parameters receive the manual C flag for this file."
                            )
                        else:
                            st.info("No mapped data parameters available for caution selection in this dataset.")

                    # Save Config
                    # Build the combined list of visit windows:
                    #   - auto_visit_windows: from MetadataLog (already validated against file range)
                    #   - manual window: from the text inputs below (if both fields filled)
                    combined_visit_windows = list(auto_visit_windows)  # copy MetadataLog windows
                    if fv_in and fv_out:
                        combined_visit_windows.append((fv_in, fv_out))  # add manual window

                    processed_file_configs.append({
                        "file": file,
                        "data_id": data_id,
                        "meta": {
                            "Logger_ID": logger_id,
                            "Logger_Script": logger_script,
                            "Logger_Software": logger_soft
                        },
                        "visit_windows": combined_visit_windows,  # list of (in, out) tuples
                        "mapping": final_mapping,
                        "add_caution": add_caution,
                        "caution_columns": selected_caution_columns,
                        "skip_rows": skip_rows,  # user-configured row skip list
                    })


            # --- Processing Button ---
            if st.button("Process & Concatenate Datasets", type="primary"):
                with st.spinner("Processing Files..."):
                    all_dfs = []
                    
                    for cfg in processed_file_configs:
                        df = process_file_data(
                            cfg['file'],
                            cfg['mapping'],
                            cfg['meta'],
                            cfg['data_id'],
                            station_name,
                            skip_rows=cfg.get('skip_rows', [0, 2, 3])  # use per-file skip_rows
                        )
                        
                        if not df.empty:
                            # Apply Field Visits Here? Or after concat? Only matters for flagging.
                            # Better to store field visits and apply to final df
                            # But we need to know which file/period the visit applies to if we do it globally? 
                            # Actually, field visits are specific time ranges, so we can apply globally by range.
                            # Actually, field visits are specific time ranges, so we can apply globally by range.
                            
                            # Apply Manual Caution Flag if selected
                            if cfg.get('add_caution', False):
                                caution_cols = cfg.get('caution_columns')
                                if caution_cols is None:
                                    caution_cols = sorted({
                                        canonicalize_column_name(str(v).strip())
                                        for v in cfg.get('mapping', {}).values()
                                        if v not in [None, "", "REMOVE"]
                                        and str(v).strip() not in [
                                            'TIMESTAMP', 'UTC Offset', 'RECORD',
                                            'Data_ID', 'Station_ID', 'Logger_ID',
                                            'Logger_Script', 'Logger_Software'
                                        ]
                                        and not str(v).strip().endswith('_Flag')
                                    })
                                for col in caution_cols:
                                    if col in df.columns:
                                        flag_col = f"{col}_Flag"
                                        if flag_col not in df.columns:
                                            df[flag_col] = "" # Init
                                        
                                        # Append "C"
                                        # Append "C" ensuring no duplicates
                                        curr = df[flag_col].fillna("").astype(str)
                                        # If "C" is already in the flag (e.g. "C", "C, T", "M, C"), don't add
                                        mask_has_c = curr.str.contains(r'\bC\b', regex=True)
                                        df.loc[~mask_has_c, flag_col] = np.where(curr[~mask_has_c] == "", "C", curr[~mask_has_c] + ", C")
                                        
                        all_dfs.append(df)
                    
                    if all_dfs:
                        # Concatenate
                        full_df = pd.concat(all_dfs, ignore_index=True)
                        
                        # Time processing
                        if 'TIMESTAMP' in full_df.columns:
                            # Stable sort keeps deterministic precedence when overlapping
                            # timestamps exist across multiple source files.
                            full_df = full_df.sort_values('TIMESTAMP', kind='mergesort')
                            full_df = full_df.drop_duplicates(subset=['TIMESTAMP'], keep='first')
                            full_df = full_df.set_index('TIMESTAMP')
                            
                            # Resample 15T
                            full_df = full_df.resample('15T').asfreq()
                            
                            # Determine flags for Resampled
                            # Meta cols needed for rebuild
                            meta_cols = ['Data_ID', 'Station_ID', 'Logger_ID', 'Logger_Script', 'Logger_Software']
                            
                            # Reset index to get Timestamp back
                            df_final = full_df.reset_index()
                            
                            # Fill Metadata
                            # Station ID is constant
                            df_final['Station_ID'] = df_final['Station_ID'].fillna(station_name)
                            if 'UTC Offset' in df_final.columns:
                                df_final['UTC Offset'] = df_final['UTC Offset'].fillna('-07:00')
                            # Others ffill/bfill
                            for mc in meta_cols:
                                if mc in df_final.columns:
                                    df_final[mc] = df_final[mc].ffill().bfill()
                            
                            # Flags Logic
                            # 1. Create Flags
                            # 2. Check ERR (Infinity/NaN conversion)
                            # 3. Check M (Missing)
                            # 4. Check V (Field Visits)
                            
                            data_cols = [
                                c for c in df_final.columns
                                if c not in meta_cols and c not in ['TIMESTAMP', 'UTC Offset', 'RECORD']
                            ]
                            
                            # Collect all field visit windows from all uploaded files.
                            # Each entry is either:
                            #   (datetime, datetime)  — from MetadataLog auto-parse
                            #   (str, str)            — from manual text input
                            all_field_visits = []
                            for cfg in processed_file_configs:
                                for window in cfg.get('visit_windows', []):
                                    all_field_visits.append(window)

                            for col in data_cols:
                                # Skip flag columns and RECORD (handled separately)
                                if col.endswith('_Flag') or col == "RECORD":
                                    continue
                                # Timestamp-like columns should not have flag columns
                                # (same treatment as TIMESTAMP).
                                if col in TIMESTAMP_LIKE_COLUMNS:
                                    df_final[col] = pd.to_datetime(df_final[col], errors='coerce')
                                    ts_flag_col = f"{col}_Flag"
                                    if ts_flag_col in df_final.columns:
                                        df_final = df_final.drop(columns=[ts_flag_col])
                                    continue

                                flag_col = f"{col}_Flag"
                                # Don't overwrite existing flags (e.g. C from file processing)
                                if flag_col not in df_final.columns:
                                    df_final[flag_col] = "" # Init
                                else:
                                    df_final[flag_col] = df_final[flag_col].fillna("").astype(str)
                                
                                # ERR Logic (numeric sensor channels only).
                                original = df_final[col].copy()
                                numeric = pd.to_numeric(df_final[col], errors='coerce')
                                
                                # Check for corruption (was present but became NaN) or Infinity
                                was_present = original.notna() & (original != "")
                                became_nan = numeric.isna()
                                is_inf = np.isinf(numeric)
                                
                                mask_err = is_inf | (was_present & became_nan)
                                
                                if mask_err.any():
                                    df_final.loc[mask_err, flag_col] = "E"
                                    numeric[mask_err] = np.nan
                                    
                                df_final[col] = numeric
                                
                                # M Logic (Missing)
                                mask_missing = df_final[col].isna()
                                # If missing and NOT E -> M
                                has_e = df_final[flag_col].fillna("").astype(str).str.contains(r'\bE\b', regex=True)
                                mask_m = mask_missing & ~has_e
                                df_final.loc[mask_m, flag_col] = "M"
                                
                                # V Logic (Field Visits)
                                # Handles both datetime objects (from MetadataLog) and strings (manual input).
                                for f_in, f_out in all_field_visits:
                                    try:
                                        # Convert to datetime if they came in as strings
                                        dt_in  = f_in  if isinstance(f_in,  datetime) else pd.to_datetime(f_in)
                                        dt_out = f_out if isinstance(f_out, datetime) else pd.to_datetime(f_out)

                                        # Round in DOWN to nearest 15-min interval
                                        t_start = pd.Timestamp(dt_in).floor('15min')
                                        # Round out UP to next 15-min interval
                                        t_end   = pd.Timestamp(dt_out).ceil('15min')

                                        mask_visit = (
                                            (df_final['TIMESTAMP'] >= t_start) &
                                            (df_final['TIMESTAMP'] <= t_end)
                                        )

                                        # Append V flag (avoid duplicates)
                                        current_flags = df_final.loc[mask_visit, flag_col].fillna("").astype(str)
                                        # If E is present, keep E-only semantics (do not append V).
                                        new_flags = np.where(
                                            current_flags.str.contains(r'\bE\b|\bV\b', regex=True),
                                            current_flags,
                                            np.where(current_flags == "", "V", current_flags + ", V")
                                        )
                                        df_final.loc[mask_visit, flag_col] = new_flags

                                    except Exception as e:
                                        st.warning(f"Invalid Field Visit Time: {f_in} – {f_out}: {e}")

                            # Ensure timestamp-like TMx fields never carry flag columns.
                            for ts_col in TIMESTAMP_LIKE_COLUMNS:
                                ts_fc = f"{ts_col}_Flag"
                                if ts_fc in df_final.columns:
                                    df_final = df_final.drop(columns=[ts_fc])

                            # RECORD missingness should also be marked M.
                            if 'RECORD' in df_final.columns:
                                rec_flag_col = 'RECORD_Flag'
                                if rec_flag_col not in df_final.columns:
                                    df_final[rec_flag_col] = ""
                                else:
                                    df_final[rec_flag_col] = df_final[rec_flag_col].fillna("").astype(str)

                                rec_vals = pd.to_numeric(df_final['RECORD'], errors='coerce')
                                rec_curr = df_final[rec_flag_col].fillna("").astype(str)
                                rec_missing = rec_vals.isna()
                                rec_has_e = rec_curr.str.contains(r'\bE\b', regex=True)
                                rec_has_m = rec_curr.str.contains(r'\bM\b', regex=True)
                                rec_apply = rec_missing & ~rec_has_e & ~rec_has_m
                                if rec_apply.any():
                                    df_final.loc[rec_apply, rec_flag_col] = np.where(
                                        rec_curr.loc[rec_apply] == "",
                                        "M",
                                        rec_curr.loc[rec_apply] + ", M"
                                    )


                            # Reorder Columns
                            # Interleave Data and Flags
                            ordered_cols = ['TIMESTAMP']
                            if 'UTC Offset' in df_final.columns:
                                ordered_cols.append('UTC Offset')
                            
                            # Handle RECORD and RECORD_Flag
                            if 'RECORD' in df_final.columns:
                                ordered_cols.append('RECORD')
                                if 'RECORD_Flag' not in df_final.columns:
                                    df_final['RECORD_Flag'] = ""
                                ordered_cols.append('RECORD_Flag')

                            # Identify data columns (exclude reserved)
                            # Reserved: TIMESTAMP, UTC Offset, RECORD, RECORD_Flag, Meta Cols, and ALL Flag columns
                            reserved = set(['TIMESTAMP', 'UTC Offset', 'RECORD', 'RECORD_Flag']) | set(meta_cols) | set([c for c in df_final.columns if c.endswith("_Flag")])
                            data_cols = [c for c in df_final.columns if c not in reserved]

                            # Add data columns and their flags
                            for col in data_cols:
                                ordered_cols.append(col)
                                flag_col = f"{col}_Flag"
                                if flag_col in df_final.columns:
                                    ordered_cols.append(flag_col)
                                        
                            # Add metadata columns at the END
                            for mc in meta_cols:
                                if mc in df_final.columns:
                                    ordered_cols.append(mc)
                            # Just ensure all columns are covered or dropped?
                            # For now, this covers the main requirement.
                            
                            df_final = df_final[ordered_cols]
                            
                            # Save
                            filename = f"{station_name}_concatenated_tidy.csv"
                            save_path = os.path.join(output_dir, filename)
                            
                            # Use helper to include units row
                            write_csv_with_units(df_final, save_path)
                            
                            st.success(f"Successfully processed {len(df_final)} records!")
                            st.success(f"Saved to: {save_path}")
                            st.dataframe(df_final.head(50))
                            
                        else:
                            st.error("Concatenation failed: Missing TIMESTAMP column.")


    # --- Tab 2: QA/QC ---
    with tab2:
        st.header("2. QA/QC Processing")

        # File Selection (Moved Up)
        if os.path.exists(output_dir):
            files = [f for f in os.listdir(output_dir) if f.endswith("_concatenated_tidy.csv")]
        else:
            files = []
        selected_file = st.selectbox("Select File to Process", files)
        
        # Load File Metadata (Dates) if selected
        file_start_date = datetime.today().date()
        file_end_date = datetime.today().date()
        
        if selected_file:
            file_path = os.path.join(output_dir, selected_file)
            try:
                # Read timestamp column, skip units row (row index 1)
                df_dates = pd.read_csv(file_path, usecols=['TIMESTAMP'], skiprows=[1])
                df_dates['TIMESTAMP'] = pd.to_datetime(df_dates['TIMESTAMP'])
                file_start_date = df_dates['TIMESTAMP'].min().date()
                file_end_date = df_dates['TIMESTAMP'].max().date()
                st.success(f"📅 File date range: {file_start_date} to {file_end_date}")
            except Exception as e:
                st.warning(f"Could not load file dates: {e}. Using today's date as default.")

        st.divider()

        # --- Instrument Configuration ---
        with st.expander("🛠️ Configure Instruments & Deployments", expanded=True):
            tab_groups, tab_deploy = st.tabs(["Instrument Groups", "Deployment History"])
            
            # --- Tab A: Instrument Groups ---
            with tab_groups:
                st.info("Define sets of instruments (e.g., 'ClimaVue50', 'Winter Setup') and their specific thresholds.")
                
                # Load
                groups = load_instrument_groups()
                group_names = list(groups.keys())
                
                # Edit / Create
                col_grp1, col_grp2 = st.columns([1, 2])
                with col_grp1:
                    selected_group = st.selectbox("Select Group to Edit", ["<New Group>"] + group_names)
                    
                    if selected_group == "<New Group>":
                        new_grp_name = st.text_input("New Group Name")
                        grp_name = new_grp_name if new_grp_name else None
                        current_cols = []
                        current_thresholds = {}
                        current_sensor_height = 200
                    else:
                        grp_name = selected_group
                        current_data = groups[selected_group]
                        # Handle both old and new format
                        if isinstance(current_data, dict) and "thresholds" in current_data:
                            current_thresholds = current_data["thresholds"]
                            current_sensor_height = current_data.get("sensor_height", 200)
                        else:
                            # Legacy format: dict is just thresholds
                            current_thresholds = current_data
                            current_sensor_height = 200
                        current_cols = list(current_thresholds.keys())

                with col_grp2:
                    if grp_name:
                        # Sensor Height Input
                        grp_sensor_height = st.number_input(
                            "Sensor Height (cm)", 
                            value=current_sensor_height, 
                            min_value=0, 
                            max_value=500,
                            key=f"sensor_height_{grp_name}"
                        )
                        st.caption("Used for sensor-height limits on DT_Avg, DBTCDT_Avg, and SV_DBTCDT_Avg.")
                        st.divider()
                        # Select Columns
                        # Get all available columns from mapping for suggestions
                        mapping = load_mapping()
                        all_known_cols = [canonicalize_column_name(c) for c in (list(mapping.keys()) + list(DEFAULT_THRESHOLDS.keys()))]
                        all_known_cols = sorted(list(set(all_known_cols)))
                        
                        # Multiselect
                        # Pre-select columns that are in the current group
                        default_sel = [c for c in current_cols if c in all_known_cols]
                        # Ensure we don't lose cols that might not be in "known"
                        extras = [c for c in current_cols if c not in all_known_cols]
                        
                        selected_cols = st.multiselect("Included Columns", all_known_cols + extras, default=default_sel + extras)
                        selected_cols = list(dict.fromkeys([canonicalize_column_name(c) for c in selected_cols]))
                        
                        # Threshold Editor for Selected Cols
                        if selected_cols:
                            st.caption("Set Thresholds for this Group (R = hard limit, C = soft caution):")
                            edit_data = []
                            for c in selected_cols:
                                # Extract R/C limits from new dict format, legacy [min,max], or defaults
                                if c in current_thresholds:
                                    val = current_thresholds[c]
                                    if isinstance(val, dict):
                                        cur_r_min = val.get('r_min') if val.get('r_min') is not None else ''
                                        cur_r_max = val.get('r_max') if val.get('r_max') is not None else ''
                                        cur_c_min = val.get('c_min') if val.get('c_min') is not None else ''
                                        cur_c_max = val.get('c_max') if val.get('c_max') is not None else ''
                                    elif isinstance(val, (list, tuple)) and len(val) >= 2:
                                        cur_r_min, cur_r_max = val[0], val[1]
                                        cur_c_min, cur_c_max = '', ''
                                    else:
                                        cur_r_min, cur_r_max, cur_c_min, cur_c_max = '', '', '', ''
                                elif c in SENSOR_THRESHOLDS:
                                    spec = SENSOR_THRESHOLDS[c]
                                    cur_r_min = spec.get('r_min') if spec.get('r_min') is not None else ''
                                    cur_r_max = spec.get('r_max') if spec.get('r_max') is not None else ''
                                    cur_c_min = spec.get('c_min') if spec.get('c_min') is not None else ''
                                    cur_c_max = spec.get('c_max') if spec.get('c_max') is not None else ''
                                else:
                                    cur_r_min, cur_r_max, cur_c_min, cur_c_max = '', '', '', ''

                                # Resolve formula tokens for UI display (e.g. H-50 -> numeric).
                                cur_r_min = resolve_height_formula_token(cur_r_min, grp_sensor_height)
                                cur_r_max = resolve_height_formula_token(cur_r_max, grp_sensor_height)
                                cur_c_min = resolve_height_formula_token(cur_c_min, grp_sensor_height)
                                cur_c_max = resolve_height_formula_token(cur_c_max, grp_sensor_height)

                                # If DBTCDT_Avg max was previously saved as blank/null, restore
                                # the expected sensor-height-dependent default in the editor.
                                if c == 'DBTCDT_Avg' and (
                                    cur_r_max == '' or cur_r_max is None or
                                    (isinstance(cur_r_max, float) and np.isnan(cur_r_max))
                                ):
                                    cur_r_max = float(grp_sensor_height) - 50.0
                                # Same behavior for SnowVue depth (H-40).
                                if c == 'SV_DBTCDT_Avg' and (
                                    cur_r_max == '' or cur_r_max is None or
                                    (isinstance(cur_r_max, float) and np.isnan(cur_r_max))
                                ):
                                    cur_r_max = float(grp_sensor_height) - 40.0
                                
                                edit_data.append({
                                    "Column": c,
                                    "R Min": cur_r_min, "R Max": cur_r_max,
                                    "C Min": cur_c_min, "C Max": cur_c_max,
                                })
                            
                            grp_df = pd.DataFrame(edit_data)
                            
                            # Configure data editor column hints.
                            column_config = {
                                "R Max": st.column_config.NumberColumn(
                                    "R Max",
                                    help="Height-based defaults: DBTCDT_Avg = H-50, SV_DBTCDT_Avg = H-40.",
                                )
                            }
                            
                            edited_grp_df = st.data_editor(
                                grp_df, 
                                key=f"editor_{grp_name}_{grp_sensor_height}", 
                                use_container_width=True,
                                column_config=column_config,
                                disabled=["Column"]  # Prevent column name editing
                            )
                            
                            if st.button("Save Group"):
                                new_thresholds = {}
                                for idx, row in edited_grp_df.iterrows():
                                    # Convert empty strings back to None for the R/C dict
                                    def _to_num_or_none(x):
                                        x = resolve_height_formula_token(x, grp_sensor_height)
                                        if x == '' or x is None or (isinstance(x, float) and np.isnan(x)):
                                            return None
                                        try:
                                            return float(x)
                                        except (ValueError, TypeError):
                                            return None
                                    r_max_val = _to_num_or_none(row.get('R Max'))
                                    # Keep DBTCDT_Avg physically bounded even when left blank.
                                    if row['Column'] == 'DBTCDT_Avg' and r_max_val is None:
                                        r_max_val = float(grp_sensor_height) - 50.0
                                    # Keep SV_DBTCDT_Avg physically bounded even when left blank.
                                    if row['Column'] == 'SV_DBTCDT_Avg' and r_max_val is None:
                                        r_max_val = float(grp_sensor_height) - 40.0
                                    new_thresholds[row['Column']] = {
                                        'r_min': _to_num_or_none(row.get('R Min')),
                                        'r_max': r_max_val,
                                        'c_min': _to_num_or_none(row.get('C Min')),
                                        'c_max': _to_num_or_none(row.get('C Max')),
                                    }
                                
                                # Save with new structure
                                groups[grp_name] = {
                                    "sensor_height": grp_sensor_height,
                                    "thresholds": new_thresholds
                                }
                                save_instrument_groups(groups)
                                st.success(f"Saved group '{grp_name}'!")
                                st.rerun()

            # --- Tab B: Deployment History ---
            with tab_deploy:
                st.info(f"Assign Instrument Groups to specific time ranges for Station: **{station_name}**")
                
                configs = load_station_configs()
                
                # Get station data (new format only)
                station_data = configs.get(station_name, {"latitude": 53.7217, "longitude": -125.6417, "deployments": []})
                st_cfg = station_data.get("deployments", [])
                
                # Display Current Configs
                if st_cfg:
                    st.write("Current Deployments:")
                    cfg_df = pd.DataFrame(st_cfg)
                    st.dataframe(cfg_df)
                    
                    if st.button("Clear History"):
                        station_data["deployments"] = []
                        configs[station_name] = station_data
                        save_station_configs(configs)
                        st.rerun()
                else:
                    st.warning("No deployment history found. 'Base' thresholds will apply everywhere.")

                st.divider()
                st.write("Add Deployment:")

                # Build a list of time options in 15-minute steps (00:00 → 23:45)
                time_options = [
                    f"{h:02d}:{m:02d}"
                    for h in range(24)
                    for m in (0, 15, 30, 45)
                ]

                c1, c2, c3 = st.columns(3)
                with c1:
                    d_start = st.date_input("Start Date", value=file_start_date, key="dep_start_date")
                    t_start_str = st.selectbox(
                        "Start Time (15-min)",
                        time_options,
                        index=0,           # defaults to 00:00
                        key="dep_start_time"
                    )
                with c2:
                    d_end = st.date_input("End Date", value=file_end_date, key="dep_end_date")
                    t_end_str = st.selectbox(
                        "End Time (15-min)",
                        time_options,
                        index=len(time_options) - 1,   # defaults to 23:45
                        key="dep_end_time"
                    )
                with c3:
                    d_grp = st.selectbox("Instrument Group", group_names)

                if st.button("Add Assignment"):
                    # Combine date + time into full datetime strings
                    start_dt_str = f"{d_start} {t_start_str}:00"
                    end_dt_str   = f"{d_end} {t_end_str}:00"

                    if start_dt_str >= end_dt_str:
                        st.error("Start datetime must be before end datetime.")
                    else:
                        # Note: lat/lon are now stored at station level, not per deployment
                        new_entry = {
                            "start": start_dt_str,   # e.g. "2024-06-01 08:15:00"
                            "end":   end_dt_str,     # e.g. "2024-09-30 23:45:00"
                            "group": d_grp
                        }
                        if station_name not in configs:
                            configs[station_name] = {"latitude": latitude, "longitude": longitude, "deployments": []}

                        configs[station_name]["deployments"].append(new_entry)
                        # Sort by start datetime string (ISO format sorts lexicographically)
                        configs[station_name]["deployments"].sort(key=lambda x: x['start'])
                        save_station_configs(configs)
                        st.success("Added deployment!")
                        st.rerun()
                
                st.divider()
                st.subheader("🔍 Check Active Thresholds")

                # Build 15-min time options for the preview picker as well
                _preview_time_opts = [
                    f"{h:02d}:{m:02d}"
                    for h in range(24)
                    for m in (0, 15, 30, 45)
                ]
                _check_col1, _check_col2 = st.columns(2)
                with _check_col1:
                    check_date = st.date_input("Preview date:", value=file_start_date, key="preview_date")
                with _check_col2:
                    check_time_str = st.selectbox(
                        "Preview time (15-min):",
                        _preview_time_opts,
                        index=0,
                        key="preview_time"
                    )
                check_dt_str = f"{check_date} {check_time_str}:00"

                # Logic to find active deployments/groups at the selected timestamp
                check_dt = pd.to_datetime(check_dt_str)

                def _extract_group_payload(group_name):
                    grp_data = groups.get(group_name, {})
                    if isinstance(grp_data, dict) and "thresholds" in grp_data:
                        return grp_data.get("thresholds", {}), grp_data.get("sensor_height", 200)
                    if isinstance(grp_data, dict):
                        return grp_data, 200
                    return {}, 200

                active_deps = []
                for cfg in st_cfg:
                    try:
                        cfg_start = pd.to_datetime(cfg.get('start'))
                        cfg_end = pd.to_datetime(cfg.get('end'))
                    except Exception:
                        continue
                    if cfg_start <= check_dt <= cfg_end:
                        active_deps.append(cfg)

                active_group_names = []
                for dep in active_deps:
                    g_name = dep.get('group')
                    if g_name and g_name not in active_group_names:
                        active_group_names.append(g_name)

                if active_group_names:
                    st.write(f"**Active Groups:** {', '.join(active_group_names)}")
                    height_labels = []
                    for g_name in active_group_names:
                        _, g_height = _extract_group_payload(g_name)
                        height_labels.append(f"{g_name}: {g_height} cm")
                    st.write(f"**Sensor Heights by Group:** {' | '.join(height_labels)}")
                else:
                    st.write("**Active Groups:** None (Using Defaults)")
                    st.write("**Sensor Height:** 200 cm (default)")
                
                # Build Comparison Table — show R and C limits side by side
                preview_data = []
                for k, v in SENSOR_THRESHOLDS.items():
                    # Default from SENSOR_THRESHOLDS
                    def_r_min = v.get('r_min', '')
                    def_r_max = v.get('r_max', '')
                    def_c_min = v.get('c_min', '')
                    def_c_max = v.get('c_max', '')
                    default_sensor_height = 200
                    
                    # Override from all active instrument groups (last matching deployment wins)
                    act_r_min = resolve_height_formula_token(def_r_min, default_sensor_height)
                    act_r_max = resolve_height_formula_token(def_r_max, default_sensor_height)
                    act_c_min = resolve_height_formula_token(def_c_min, default_sensor_height)
                    act_c_max = resolve_height_formula_token(def_c_max, default_sensor_height)
                    source = "Default"
                    source_group = ""

                    for dep in active_deps:
                        dep_group = dep.get('group')
                        grp_thresholds, grp_sensor_height = _extract_group_payload(dep_group)
                        grp_spec, _matched_key = get_threshold_spec_for_column(grp_thresholds, k)
                        if grp_spec is None:
                            continue
                        if isinstance(grp_spec, dict):
                            act_r_min = resolve_height_formula_token(grp_spec.get('r_min', def_r_min), grp_sensor_height)
                            act_r_max = resolve_height_formula_token(grp_spec.get('r_max', def_r_max), grp_sensor_height)
                            act_c_min = resolve_height_formula_token(grp_spec.get('c_min', def_c_min), grp_sensor_height)
                            act_c_max = resolve_height_formula_token(grp_spec.get('c_max', def_c_max), grp_sensor_height)
                            if k == 'DBTCDT_Avg' and act_r_max is None:
                                act_r_max = float(grp_sensor_height) - 50.0
                            if k == 'SV_DBTCDT_Avg' and act_r_max is None:
                                act_r_max = float(grp_sensor_height) - 40.0
                        elif isinstance(grp_spec, (list, tuple)) and len(grp_spec) >= 2:
                            act_r_min, act_r_max = grp_spec[0], grp_spec[1]
                            act_c_min, act_c_max = def_c_min, def_c_max
                        source = "Instrument Group"
                        source_group = dep_group
                    
                    # Format None as '—'
                    fmt = lambda x: '—' if x is None or x == '' or (isinstance(x, float) and np.isnan(x)) else str(x)
                    preview_data.append({
                        "Column": k,
                        "R Min": fmt(act_r_min), "R Max": fmt(act_r_max),
                        "C Min": fmt(act_c_min), "C Max": fmt(act_c_max),
                        "Source": source,
                        "Group": source_group if source_group else "—"
                    })
                
                st.dataframe(pd.DataFrame(preview_data), use_container_width=True)


        if selected_file:
            st.divider()

            # Use DEFAULT_THRESHOLDS directly as the base
            active_thresholds = DEFAULT_THRESHOLDS.copy()
            # Sensor height adjustments are now applied per-group in the QC pipeline

            # --- Logic Functions (Embedded to access st context) ---
            def run_qc_pipeline(df):
                # Load Time-Varying Configs
                instr_groups = load_instrument_groups()
                st_configs = load_station_configs()
                
                # Get station data (new format only)
                station_data = st_configs.get(station_name, {"latitude": 53.7217, "longitude": -125.6417, "deployments": []})
                current_deps = station_data.get("deployments", [])
                station_lat = station_data.get("latitude", 53.7217)
                station_lon = station_data.get("longitude", -125.6417)

                
                # 1. Apply Thresholds — dual-tier R (hard) and C (soft/caution)
                # ---------------------------------------------------------------
                # For each data column we:
                #   a) Look up r_min/r_max from SENSOR_THRESHOLDS (hard limits → flag R)
                #   b) Look up c_min/c_max from SENSOR_THRESHOLDS (soft limits → flag C)
                #   c) Apply time-varying overrides from instrument-group deployments
                #      (deployments store a single min/max pair used as the R limit)
                #   d) Skip a tier entirely if both its limits are None
                # ---------------------------------------------------------------

                # Helper: resolve a limit value that may be a column reference or sentinel string
                def resolve_limit(v, data_slice, sensor_height):
                    """
                    Resolves a threshold value to a numeric scalar or Series.
                    Handles:
                      - None          → returns None (skip this limit)
                      - 'H+5'         → sensor_height + 5
                      - 'H-50'        → sensor_height - 50
                      - column name   → numeric Series from that column
                      - numeric       → returned as-is
                    """
                    if v is None:
                        return None
                    if isinstance(v, str):
                        # Handle H+N or H-N patterns (e.g. H-50, H-40, H+5)
                        hm = re.match(r'^H([+-]\d+(?:\.\d+)?)$', v)
                        if hm:
                            return sensor_height + float(hm.group(1))
                        if v in data_slice.columns:
                            return pd.to_numeric(data_slice[v], errors='coerce')
                    return v

                # Helper: append a flag token to a flag column, skipping rows already flagged M/ERR/E
                def _append_flag(df, flag_col, mask, token):
                    """Appends token to flag_col for rows where mask is True, skipping M/ERR/E rows."""
                    if not mask.any():
                        return
                    curr = df[flag_col].fillna("").astype(str)
                    # Don't overwrite M, ERR, or E rows.
                    skip = curr.str.contains(r'\bM\b|\bERR\b|\bE\b', regex=True)
                    apply_mask = mask & ~skip
                    if not apply_mask.any():
                        return
                    targets = curr.loc[apply_mask]
                    pat = rf'\b{token}\b'
                    already = targets.str.contains(pat, regex=True)
                    new_flags = np.where(
                        already,
                        targets,
                        np.where(targets == "", token, targets + ", " + token)
                    )
                    df.loc[apply_mask, flag_col] = new_flags

                def _normalize_flag_cell(val):
                    """
                    Normalize a flag-cell string by removing blanks/nan tokens and
                    de-duplicating tokens while preserving first-seen order.
                    """
                    if val is None:
                        return ""
                    if isinstance(val, float) and np.isnan(val):
                        return ""
                    text = str(val).strip()
                    if text == "" or text.lower() in ["nan", "none"]:
                        return ""

                    seen = set()
                    ordered = []
                    for token in text.split(","):
                        t = token.strip()
                        if not t or t.lower() in ["nan", "none"]:
                            continue
                        if t in seen:
                            continue
                        seen.add(t)
                        ordered.append(t)
                    return ", ".join(ordered)

                def _dedupe_all_flag_columns(df):
                    for fc in [c for c in df.columns if c.endswith("_Flag")]:
                        df[fc] = df[fc].apply(_normalize_flag_cell)

                def _resolve_dep_col(df, col_name):
                    """
                    Resolve a dependency column name to an existing DataFrame column,
                    checking known threshold/dependency aliases (e.g., MaxWS_ms <-> MaxWS_ms_Avg).
                    """
                    for candidate in threshold_key_variants(col_name):
                        if candidate in df.columns:
                            return candidate
                    return None

                # Helper: extract thresholds dict from an instrument-group entry
                def _get_grp_thresholds(grp_data):
                    if isinstance(grp_data, dict) and "thresholds" in grp_data:
                        return grp_data["thresholds"], grp_data.get("sensor_height", 200)
                    return grp_data, 200

                # Identify columns to QC
                qc_cols = [
                    c for c in df.columns
                    if not c.endswith('_Flag')
                    and c not in ['TIMESTAMP', 'UTC Offset', 'RECORD', 'Data_ID', 'Station_ID',
                                  'Logger_ID', 'Logger_Script', 'Logger_Software']
                ]

                for col in qc_cols:
                    # Skip timestamp-like columns (no numeric thresholds apply)
                    if col in TIMESTAMP_LIKE_COLUMNS:
                        df[col] = pd.to_datetime(df[col], errors='coerce')
                        ts_flag_col = f"{col}_Flag"
                        if ts_flag_col in df.columns:
                            df = df.drop(columns=[ts_flag_col])
                        continue

                    # --- Determine base R and C limits from SENSOR_THRESHOLDS ---
                    base_spec = SENSOR_THRESHOLDS.get(col)

                    # If column not in SENSOR_THRESHOLDS, check if any deployment group covers it
                    if base_spec is None:
                        has_dep_coverage = any(
                            get_threshold_spec_for_column(
                                _get_grp_thresholds(instr_groups.get(d['group'], {}))[0],
                                col
                            )[0] is not None
                            for d in current_deps
                        )
                        if not has_dep_coverage:
                            continue  # no threshold info at all — skip
                        # Deployment group covers it but no global spec: use inf as base
                        base_r_min, base_r_max = -np.inf, np.inf
                        base_c_min, base_c_max = None, None
                    else:
                        base_r_min = base_spec.get('r_min')
                        base_r_max = base_spec.get('r_max')
                        base_c_min = base_spec.get('c_min')
                        base_c_max = base_spec.get('c_max')

                    # Ensure flag column exists
                    flag_col = f"{col}_Flag"
                    if flag_col not in df.columns:
                        df[flag_col] = ""

                    vals = pd.to_numeric(df[col], errors='coerce')

                    # --- Build time-varying limit Series for R tier ---
                    # Default sensor height (overridden per deployment below)
                    default_sensor_height = 200

                    # Check if any deployment overrides this column
                    relevant_deps = [
                        d for d in current_deps
                        if get_threshold_spec_for_column(
                            _get_grp_thresholds(instr_groups.get(d['group'], {}))[0],
                            col
                        )[0] is not None
                    ]

                    if not relevant_deps:
                        # No deployment override — use base limits directly (fast path)
                        r_min_eff = resolve_limit(base_r_min, df, default_sensor_height)
                        r_max_eff = resolve_limit(base_r_max, df, default_sensor_height)
                        c_min_eff = resolve_limit(base_c_min, df, default_sensor_height)
                        c_max_eff = resolve_limit(base_c_max, df, default_sensor_height)

                        # R flag (hard limit)
                        if r_min_eff is not None or r_max_eff is not None:
                            mask_r = pd.Series(False, index=df.index)
                            if r_min_eff is not None:
                                mask_r = mask_r | (vals < r_min_eff)
                            if r_max_eff is not None:
                                mask_r = mask_r | (vals > r_max_eff)
                            _append_flag(df, flag_col, mask_r, 'R')

                        # C flag (soft limit) — only for rows that did NOT get R
                        if c_min_eff is not None or c_max_eff is not None:
                            mask_already_r = df[flag_col].fillna("").str.contains(r'\bR\b', regex=True)
                            mask_c = pd.Series(False, index=df.index)
                            if c_min_eff is not None:
                                mask_c = mask_c | (vals < c_min_eff)
                            if c_max_eff is not None:
                                mask_c = mask_c | (vals > c_max_eff)
                            mask_c = mask_c & ~mask_already_r
                            _append_flag(df, flag_col, mask_c, 'C')

                    else:
                        # Time-varying: build per-row limit Series for both R and C
                        r_min_series = pd.Series(
                            resolve_limit(base_r_min, df, default_sensor_height)
                            if not isinstance(base_r_min, str) else np.nan,
                            index=df.index, dtype=float
                        )
                        r_max_series = pd.Series(
                            resolve_limit(base_r_max, df, default_sensor_height)
                            if not isinstance(base_r_max, str) else np.nan,
                            index=df.index, dtype=float
                        )
                        # C-limit series — also time-varying from group thresholds
                        c_min_series = pd.Series(
                            resolve_limit(base_c_min, df, default_sensor_height)
                            if base_c_min is not None and not isinstance(base_c_min, str) else np.nan,
                            index=df.index, dtype=float
                        )
                        c_max_series = pd.Series(
                            resolve_limit(base_c_max, df, default_sensor_height)
                            if base_c_max is not None and not isinstance(base_c_max, str) else np.nan,
                            index=df.index, dtype=float
                        )

                        for dep in current_deps:
                            grp_data = instr_groups.get(dep['group'], {})
                            grp_thresholds, grp_sensor_height = _get_grp_thresholds(grp_data)

                            col_spec, _matched_key = get_threshold_spec_for_column(grp_thresholds, col)
                            if col_spec is None:
                                continue
                            try:
                                t_s = pd.to_datetime(dep['start'])
                                t_e = pd.to_datetime(dep['end']) + timedelta(hours=23, minutes=59)
                                mask_time = (df['TIMESTAMP'] >= t_s) & (df['TIMESTAMP'] <= t_e)
                                if not mask_time.any():
                                    continue

                                # Handle both dict {r_min, r_max, c_min, c_max} and legacy [min, max]
                                if isinstance(col_spec, dict):
                                    # New R/C dict structure
                                    g_r_min = col_spec.get('r_min')
                                    g_r_max = col_spec.get('r_max')
                                    g_c_min = col_spec.get('c_min')
                                    g_c_max = col_spec.get('c_max')
                                elif isinstance(col_spec, (list, tuple)) and len(col_spec) == 2:
                                    # Legacy [min, max] list — treat as R limits only
                                    g_r_min, g_r_max = col_spec
                                    g_c_min, g_c_max = None, None
                                else:
                                    continue  # unknown format, skip

                                # Resolve R limits for this deployment window
                                g_r_min_v = resolve_limit(g_r_min, df.loc[mask_time], grp_sensor_height)
                                g_r_max_v = resolve_limit(g_r_max, df.loc[mask_time], grp_sensor_height)
                                if g_r_min_v is not None:
                                    r_min_series.loc[mask_time] = g_r_min_v
                                if g_r_max_v is not None:
                                    r_max_series.loc[mask_time] = g_r_max_v

                                # Resolve C limits for this deployment window
                                g_c_min_v = resolve_limit(g_c_min, df.loc[mask_time], grp_sensor_height)
                                g_c_max_v = resolve_limit(g_c_max, df.loc[mask_time], grp_sensor_height)
                                if g_c_min_v is not None:
                                    c_min_series.loc[mask_time] = g_c_min_v
                                if g_c_max_v is not None:
                                    c_max_series.loc[mask_time] = g_c_max_v

                            except Exception as e:
                                st.warning(f"Config Error ({dep}): {e}")

                        # R flag (hard limit, time-varying)
                        mask_r = (vals < r_min_series) | (vals > r_max_series)
                        _append_flag(df, flag_col, mask_r, 'R')

                        # C flag (soft limit, time-varying from group thresholds)
                        has_c = c_min_series.notna().any() or c_max_series.notna().any()
                        if has_c:
                            mask_already_r = df[flag_col].fillna("").str.contains(r'\bR\b', regex=True)
                            mask_c = pd.Series(False, index=df.index)
                            if c_min_series.notna().any():
                                mask_c = mask_c | (vals < c_min_series)
                            if c_max_series.notna().any():
                                mask_c = mask_c | (vals > c_max_series)
                            mask_c = mask_c & ~mask_already_r
                            _append_flag(df, flag_col, mask_c, 'C')



                # 2. Dynamic/Logic Flags
                # Snow Depth Logic - Time-varying sensor height per deployment
                if 'DBTCDT_Avg' in df.columns:
                     vals = pd.to_numeric(df['DBTCDT_Avg'], errors='coerce')
                     flag_col = 'DBTCDT_Avg_Flag'

                     # Build time-varying limit for T > H-50.
                     # Start with a default, then override only from groups that
                     # explicitly define DBTCDT_Avg thresholds (e.g., SR50).
                     default_sensor_height = 200
                     limit_series = pd.Series(default_sensor_height - 50, index=df.index)

                     for dep in current_deps:
                         grp_data = instr_groups.get(dep['group'], {})
                         grp_thresholds, grp_sensor_height = _get_grp_thresholds(grp_data)
                         col_spec, _matched_key = get_threshold_spec_for_column(grp_thresholds, 'DBTCDT_Avg')
                         if col_spec is None:
                             continue

                         try:
                             t_s = pd.to_datetime(dep['start'])
                             t_e = pd.to_datetime(dep['end']) + timedelta(hours=23, minutes=59)

                             mask_time = (df['TIMESTAMP'] >= t_s) & (df['TIMESTAMP'] <= t_e)
                             if mask_time.any():
                                 # Apply sensor height - 50 for this SR50-covered period.
                                 limit_series.loc[mask_time] = grp_sensor_height - 50
                         except Exception as e:
                             st.warning(f"Config Error in DBTCDT_Avg logic ({dep}): {e}")

                     # R > H-50 (hard limit breach — sensor-height-dependent)
                     mask_t = (vals > limit_series)
                     _append_flag(df, flag_col, mask_t, 'R')

                # SnowVue10 Snow Depth Logic — Time-varying sensor height per deployment (H-40)
                if 'SV_DBTCDT_Avg' in df.columns:
                    vals = pd.to_numeric(df['SV_DBTCDT_Avg'], errors='coerce')
                    flag_col = 'SV_DBTCDT_Avg_Flag'
                    if flag_col not in df.columns:
                        df[flag_col] = ""

                    default_sensor_height = 200
                    limit_series = pd.Series(default_sensor_height - 40, index=df.index)

                    for dep in current_deps:
                        grp_data = instr_groups.get(dep['group'], {})
                        grp_thresholds, grp_sensor_height = _get_grp_thresholds(grp_data)
                        col_spec, _matched_key = get_threshold_spec_for_column(grp_thresholds, 'SV_DBTCDT_Avg')
                        if col_spec is None:
                            continue
                        try:
                            t_s = pd.to_datetime(dep['start'])
                            t_e = pd.to_datetime(dep['end']) + timedelta(hours=23, minutes=59)
                            mask_time = (df['TIMESTAMP'] >= t_s) & (df['TIMESTAMP'] <= t_e)
                            if mask_time.any():
                                limit_series.loc[mask_time] = grp_sensor_height - 40
                        except Exception as e:
                            st.warning(f"Config Error in SV_DBTCDT_Avg logic ({dep}): {e}")

                    # R > H-40 (hard limit breach — sensor-height-dependent)
                    mask_t = (vals > limit_series)
                    _append_flag(df, flag_col, mask_t, 'R')

                # Wind logic (NV): when wind speed is 0, wind direction is invalid.
                # Apply NV to wind direction only (not WS_ms_Avg).

                if 'WS_ms_Avg' in df.columns and 'WindDir' in df.columns:
                    ws = pd.to_numeric(df['WS_ms_Avg'], errors='coerce')
                    mask_no_wind = ws.notna() & (ws == 0)
                    if mask_no_wind.any():
                        fc = 'WindDir_Flag'
                        if fc not in df.columns:
                            df[fc] = ""
                        _append_flag(df, fc, mask_no_wind, 'NV')

                        # WindDir_SD1_WVT NV when wind speed is 0
                        wsd_col = None
                        for candidate in ['WindDir_SD1_WVT', 'WindDir_SD1']:
                            if candidate in df.columns:
                                wsd_col = candidate
                                break
                        if wsd_col is not None:
                            fc_wsd = f'{wsd_col}_Flag'
                            if fc_wsd not in df.columns:
                                df[fc_wsd] = ""
                            _append_flag(df, fc_wsd, mask_no_wind, 'NV')

                        # MaxWS_ms NV when wind speed is 0
                        # Per RefSensorThresholds: "Valid only if wind > 0"
                        for maxws in ['MaxWS_ms', 'MaxWS_ms_Avg', 'MaxWS_ms_Max']:
                            if maxws in df.columns:
                                fc_mw = f'{maxws}_Flag'
                                if fc_mw not in df.columns:
                                    df[fc_mw] = ""
                                _append_flag(df, fc_mw, mask_no_wind, 'NV')
                                break  # only one variant will be present

                # Lightning-distance validity flag:
                # Dist_km_Avg is valid only when Strikes_Tot > 0.
                # Apply NV directly to Dist_km_Avg when Strikes_Tot == 0.
                if 'Strikes_Tot' in df.columns and 'Dist_km_Avg' in df.columns:
                    strikes = pd.to_numeric(df['Strikes_Tot'], errors='coerce')
                    mask_no_strikes = strikes.notna() & (strikes == 0)
                    if mask_no_strikes.any():
                        fc = 'Dist_km_Avg_Flag'
                        if fc not in df.columns:
                            df[fc] = ""
                        _append_flag(df, fc, mask_no_strikes, 'NV')

                # Note: Tilt columns (TiltNS_deg_Avg, TiltWE_deg_Avg) are handled entirely by:
                #   1. The main threshold loop above (applies C for >|3°|, R for >|90°|)
                #   2. DEPENDENCY_CONFIG (propagates T to SlrFD_W_Avg/Rain_mm_Tot when tilt has C,
                #      and DC when tilt has R)
                # No separate tilt block is needed here.

                # 3. Nighttime Flags (Z)
                if Sun and 'TIMESTAMP' in df.columns:
                    # Use station-level lat/lon (already extracted above from station config)
                    # station_lat and station_lon were set when loading station config
                    sun = Sun(station_lat, station_lon)
                    tz_pdt = timezone(timedelta(hours=-7))

                    df['TIMESTAMP'] = pd.to_datetime(df['TIMESTAMP'])
                    temp_dates = df['TIMESTAMP'].dt.date
                    unique_dates = temp_dates.unique()

                    for d in unique_dates:
                        # Find rise/set
                        rise_naive = None
                        set_naive = None
                        candidates = [datetime(d.year, d.month, d.day), datetime(d.year, d.month, d.day) + timedelta(days=1)]

                        for cand in candidates:
                            try:
                                r_utc = sun.get_sunrise_time(cand)
                                s_utc = sun.get_sunset_time(cand)
                                r_pdt = r_utc.astimezone(tz_pdt)
                                s_pdt = s_utc.astimezone(tz_pdt)

                                if r_pdt.date() == d: rise_naive = r_pdt.replace(tzinfo=None)
                                if s_pdt.date() == d: set_naive = s_pdt.replace(tzinfo=None)
                            except: continue

                        if rise_naive and set_naive:
                            mask_date = (temp_dates == d)
                            ts_vals = df.loc[mask_date, 'TIMESTAMP']
                            padding = timedelta(minutes=15)

                            mask_night = (ts_vals < (rise_naive - padding)) | (ts_vals > (set_naive + padding))
                            night_indices = ts_vals[mask_night].index

                            if len(night_indices) > 0:
                                for scol in SOLAR_COLUMNS:
                                    if scol not in df.columns:
                                        continue
                                    vals = pd.to_numeric(df.loc[night_indices, scol], errors='coerce').fillna(0)
                                    # Per RefSensorThresholds notes:
                                    #   SlrFD_W_Avg: Z when > 0 at night
                                    #   SWin_Avg/SWout_Avg: Z when < 0 at night
                                    if scol == 'SlrFD_W_Avg':
                                        mask_z = vals > 0.0001
                                    else:
                                        mask_z = vals < -0.0001
                                    if mask_z.any():
                                        idx = vals[mask_z].index
                                        fc = f"{scol}_Flag"
                                        if fc not in df.columns:
                                            df[fc] = ""
                                        mask_z_rows = pd.Series(False, index=df.index)
                                        mask_z_rows.loc[idx] = True
                                        _append_flag(df, fc, mask_z_rows, 'Z')

                # 4. E flag — sensor-specific error values (-9999 or -9990)
                # Per Notes column: E if -9999 (or -9990 for WS_ms_Avg)
                # These are logger-encoded error codes that indicate sensor failure.
                # Must run BEFORE BV/PT propagation so that error-value rows have
                # their flags correctly set to E before system-level propagation.
                ERROR_VALUES = {-9999, -9990, -9998}
                for col in qc_cols:
                    if col not in df.columns:
                        continue
                    if col in TIMESTAMP_LIKE_COLUMNS:
                        continue
                    flag_col = f"{col}_Flag"
                    if flag_col not in df.columns:
                        df[flag_col] = ""
                    raw_vals = pd.to_numeric(df[col], errors='coerce')
                    mask_err_val = raw_vals.isin(ERROR_VALUES)
                    # Per RefSensorThresholds notes for DT: "E if 0 (no echo detected)"
                    if col in ('DT_Avg', 'SV_DT_Avg'):
                        mask_err_val = mask_err_val | raw_vals.eq(0)
                    # Per RefSensorThresholds: SV_Q_Avg "E if 0"
                    if col == 'SV_Q_Avg':
                        mask_err_val = mask_err_val | raw_vals.eq(0)
                    # Per RefSensorThresholds: SV_Alert "E if 1" (alert active)
                    if col == 'SV_Alert':
                        mask_err_val = mask_err_val | raw_vals.eq(1)
                    if mask_err_val.any():
                        # E is exclusive: if present, suppress all other flags in this cell.
                        df.loc[mask_err_val, flag_col] = "E"

                # 5. System-level propagation: BV (battery voltage R) and PT (panel temp R)
                # When BattV_Avg or PTemp_C_Avg is flagged R, all other sensor columns
                # get BV or PT appended to their flag (per Flags_Depend in RefSensorThresholds).
                # This is done programmatically because it affects every column.
                system_propagations = [
                    ('BattV_Avg', 'R', 'BV'),
                    ('PTemp_C_Avg', 'R', 'PT'),
                    ('Ptmp_C_Avg', 'R', 'PT'),
                ]
                for src_col, trigger_flag, prop_flag in system_propagations:
                    src_fc = f"{src_col}_Flag"
                    if src_fc not in df.columns:
                        continue
                    # Find rows where the source column has the trigger flag
                    mask_src = df[src_fc].fillna("").astype(str).str.contains(rf'\b{trigger_flag}\b', regex=True)
                    if not mask_src.any():
                        continue
                    # Propagate to all other sensor columns (skip metadata, TIMESTAMP, RECORD, and the source itself)
                    skip_cols = {'TIMESTAMP', 'RECORD', src_col, src_fc}
                    for col in df.columns:
                        if col.endswith('_Flag') and col not in skip_cols:
                            _append_flag(df, col, mask_src, prop_flag)

                # 6. Critical Flags (PT — panel temperature flagged R)
                # If PTemp is flagged R, warn the user (data may be unreliable system-wide)
                for pt_col in ['PTemp_C_Avg_Flag', 'Ptmp_C_Avg_Flag']:
                    if pt_col in df.columns:
                        pf = df[pt_col].fillna("").astype(str)
                        mask_crit = pf.str.contains(r'\bR\b', regex=True)
                        if mask_crit.any():
                            st.warning(f"⚠️ Panel Temperature (PT) flagged R in {mask_crit.sum()} records — system-wide data quality may be affected.")

                # 6.5 RECORD missingness guard (M) in QA/QC stage.
                # (Numbering note: E moved to step 4, BV/PT to step 5, PT warning to step 6)
                # Keeps behavior consistent even if the source concatenated file
                # was produced before ingestion-stage RECORD M logic existed.
                if "RECORD" in df.columns:
                    rec_fc = "RECORD_Flag"
                    if rec_fc not in df.columns:
                        df[rec_fc] = ""
                    rec_vals = pd.to_numeric(df["RECORD"], errors='coerce')
                    rec_curr = df[rec_fc].fillna("").astype(str)
                    rec_missing = rec_vals.isna()
                    rec_has_e = rec_curr.str.contains(r'\bE\b', regex=True)
                    rec_has_m = rec_curr.str.contains(r'\bM\b', regex=True)
                    rec_apply = rec_missing & ~rec_has_e & ~rec_has_m
                    if rec_apply.any():
                        df.loc[rec_apply, rec_fc] = np.where(
                            rec_curr.loc[rec_apply] == "",
                            "M",
                            rec_curr.loc[rec_apply] + ", M"
                        )
                # 7. LR (Logger Restart) — RECORD sequence restart
                # Per FlagLibrary: LR indicates power failure or logger update.
                if "RECORD" in df.columns:
                     vals = pd.to_numeric(df["RECORD"], errors='coerce')
                     prev = vals.shift(1)
                     is_start = prev.isna()
                     mask_restart = (vals < prev) | (is_start & (vals==0))
                     if mask_restart.any():
                         fc = "RECORD_Flag"
                         if fc not in df.columns: df[fc] = ""
                         curr = df.loc[mask_restart, fc].fillna("").astype(str)
                         df.loc[mask_restart, fc] = np.where(curr == "", "LR", curr + ", LR")

                # 7.5 LR dependency propagation
                # LR is derived from RECORD sequence resets, then propagated to all other flags.
                if "RECORD_Flag" in df.columns:
                    mask_lr = df["RECORD_Flag"].fillna("").astype(str).str.contains(r'\bLR\b', regex=True)
                    if mask_lr.any():
                        for fc in [c for c in df.columns if c.endswith("_Flag") and c != "RECORD_Flag"]:
                            _append_flag(df, fc, mask_lr, "LR")

                # 8. DZ (Divide by Zero) — albedo when SWin < 20 W/m²
                # Per Notes: "Flag DZ if SWin_Avg < 20" (too dark to compute meaningful albedo)
                if 'SWalbedo_Avg' in df.columns and 'SWin_Avg' in df.columns:
                    sw_in = pd.to_numeric(df['SWin_Avg'], errors='coerce')
                    mask_dz = sw_in < 20
                    if mask_dz.any():
                        fc = 'SWalbedo_Avg_Flag'
                        if fc not in df.columns:
                            df[fc] = ""
                        _append_flag(df, fc, mask_dz, "DZ")

                # 9. Dependencies — propagate flags from source to target columns
                for dep in DEPENDENCY_CONFIG:
                   target = _resolve_dep_col(df, dep['target'])
                   if not target or target in TIMESTAMP_LIKE_COLUMNS:
                       continue
                   tfc = f"{target}_Flag"
                   if tfc not in df.columns:
                       continue

                   mask_fail = pd.Series(False, index=df.index)
                   for src in dep['sources']:
                       src_col = _resolve_dep_col(df, src)
                       if not src_col: continue
                       sfc = f"{src_col}_Flag"
                       if sfc not in df.columns:
                           continue
                       curr_s = df[sfc].fillna("").astype(str)
                       pat = "|".join([rf"\b{f}\b" for f in dep['trigger_flags']])
                       mask_fail = mask_fail | curr_s.str.contains(pat, regex=True)

                   if mask_fail.any():
                       _append_flag(df, tfc, mask_fail, dep['set_flag'])

                # 9.5 Normalize all flag strings
                # Prevent duplicates like "C, Z, Z" when multiple logic layers add same token.
                _dedupe_all_flag_columns(df)

                # 10. SF (Snow Free period) — apply to SR50 outputs in months 6,7,8,9.
                if 'TIMESTAMP' in df.columns:
                    months = df['TIMESTAMP'].dt.month
                    snow_free_months = [6, 7, 8, 9]
                    # Apply SF to all snow-depth workflow outputs marked V,SF in RefSensorThresholds:
                    # SR50/SR50AT: DT, Q, TCDT, snow(depth)
                    # SnowVue10:   DT, Q, TCDT, snow(depth)
                    sr50_sf_cols = [
                        'DT_Avg', 'Q_Avg', 'TCDT_Avg', 'DBTCDT_Avg',
                        'SV_DT_Avg', 'SV_Q_Avg', 'SV_TCDT_Avg', 'SV_DBTCDT_Avg'
                    ]
                    for sr_col in sr50_sf_cols:
                        if sr_col not in df.columns:
                            continue
                        fc = f'{sr_col}_Flag'
                        if fc not in df.columns:
                            df[fc] = ""
                        vals_sr = pd.to_numeric(df[sr_col], errors='coerce')
                        mask_sf = months.isin(snow_free_months) & vals_sr.notna()
                        _append_flag(df, fc, mask_sf, 'SF')

                # 11. Pass Flags (P) — clean data with no flags
                for col in df.columns:
                    if col.endswith("_Flag"):
                        dcol = col[:-5]
                        if dcol in df.columns:
                            curr_f = df[col].fillna("").astype(str).str.strip().replace('nan', '')
                            vals = df[dcol]
                            mask_p = (curr_f == "") & (vals.notna()) & (vals.astype(str).str.strip() != '') & (vals.astype(str).str.lower() != 'nan')
                            if mask_p.any():
                                df.loc[mask_p, col] = 'P'

                # Ensure timestamp-like TMx fields never carry flag columns.
                for ts_col in TIMESTAMP_LIKE_COLUMNS:
                    ts_fc = f"{ts_col}_Flag"
                    if ts_fc in df.columns:
                        df = df.drop(columns=[ts_fc])

                return df

            if st.button("Run QA/QC Pipeline", type="primary"):
                f_path = os.path.join(output_dir, selected_file)

                with st.spinner("Running QA/QC..."):
                    try:
                        # Load
                        df_qc = pd.read_csv(f_path, low_memory=False)
                        
                        # Handle Units Row (if present)
                        # Check if first row is units (e.g. TIMESTAMP is 'TS')
                        if not df_qc.empty and 'TIMESTAMP' in df_qc.columns:
                            first_val = str(df_qc.iloc[0]['TIMESTAMP'])
                            if first_val == 'TS':
                                df_qc = df_qc.iloc[1:].reset_index(drop=True)
                                
                        if 'TIMESTAMP' in df_qc.columns:
                            df_qc['TIMESTAMP'] = pd.to_datetime(df_qc['TIMESTAMP'])

                        # Process
                        df_qc = run_qc_pipeline(df_qc)

                        # Reorder Columns (Interleave)
                        # Reorder Columns (Interleave)
                        ordered_cols = ['TIMESTAMP']
                        if 'UTC Offset' in df_qc.columns:
                            ordered_cols.append('UTC Offset')
                        
                        # Handle RECORD and RECORD_Flag
                        if 'RECORD' in df_qc.columns:
                            ordered_cols.append('RECORD')
                            # Ensure flag exists (QC might have added LR, but if not, create empty)
                            if 'RECORD_Flag' not in df_qc.columns:
                                df_qc['RECORD_Flag'] = ""
                            ordered_cols.append('RECORD_Flag')
                            
                        meta_cols = ['Data_ID', 'Station_ID', 'Logger_ID', 'Logger_Script', 'Logger_Software']

                        # Identify data columns (everything else)
                        reserved = set(['TIMESTAMP', 'UTC Offset', 'RECORD', 'RECORD_Flag']) | set(meta_cols) | set([c for c in df_qc.columns if c.endswith("_Flag")])
                        data_cols = [c for c in df_qc.columns if c not in reserved]
                        
                        for col in data_cols:
                            ordered_cols.append(col)
                            flag_col = f"{col}_Flag"
                            if flag_col in df_qc.columns:
                                ordered_cols.append(flag_col)
                                
                        # Add metadata columns at the END
                        for mc in meta_cols:
                            if mc in df_qc.columns: ordered_cols.append(mc)
                                
                        df_qc = df_qc[ordered_cols]

                        # Save using SOP/FileNames convention (prefer contemporary name
                        # when logger model/serial + retrieval date can be determined).
                        out_name = build_sop_tidy_save_name(
                            df_qc=df_qc,
                            station_code=station_name,
                            output_dir=output_dir,
                            selected_input_name=selected_file
                        )
                        save_path = os.path.join(output_dir, out_name)
                        
                        # Use helper to include units row
                        write_csv_with_units(df_qc, save_path)

                        st.success("QA/QC Complete!")
                        if out_name.endswith("_tidy_historical.csv"):
                            st.info("Could not uniquely determine contemporary logger/date components; used historical SOP name.")
                        st.success(f"Saved to (SOP format): {save_path}")
                        st.dataframe(df_qc.head(50))

                    except Exception as e:
                        st.error(f"QA/QC Failed: {e}")
                        st.exception(e)

    # --- Tab 3: Visualization ---
    with tab3:
        st.header("3. QA/QC Visualization")

        if os.path.exists(output_dir):
            qc_files = sorted([f for f in os.listdir(output_dir) if is_qaqc_output_file(f)])
        else:
            qc_files = []

        if not qc_files:
            st.info("No QA/QC output files found. Run Tab 2 first to generate a tidy QA/QC file.")
        else:
            selected_qc_file = st.selectbox("Select QA/QC File to Visualize", qc_files)

            if selected_qc_file:
                qc_path = os.path.join(output_dir, selected_qc_file)
                with st.spinner("Loading QA/QC file..."):
                    df_viz, flags_long = load_qc_visualization_data(qc_path)

                st.success(f"Loaded: {qc_path}")

                info_col1, info_col2, info_col3 = st.columns(3)
                info_col1.metric("Rows", f"{len(df_viz):,}")
                info_col2.metric("Columns", f"{df_viz.shape[1]:,}")
                if 'TIMESTAMP' in df_viz.columns and df_viz['TIMESTAMP'].notna().any():
                    t_start = df_viz['TIMESTAMP'].min().strftime("%Y-%m-%d")
                    t_end = df_viz['TIMESTAMP'].max().strftime("%Y-%m-%d")
                    info_col3.metric("Date Range", f"{t_start} to {t_end}")
                else:
                    info_col3.metric("Date Range", "N/A")

                if flags_long.empty:
                    st.warning("No *_Flag columns with values were found in this file.")
                else:
                    all_flags = sorted(flags_long['flag'].unique().tolist())
                    default_flags = [f for f in all_flags if f != "P"] or all_flags
                    selected_flags = st.multiselect(
                        "Flags to include",
                        options=all_flags,
                        default=default_flags,
                        help="By default, pass flags (P) are excluded so exceptions are easier to inspect."
                    )

                    if not selected_flags:
                        st.warning("Select at least one flag to visualize.")
                    else:
                        filtered = flags_long[flags_long['flag'].isin(selected_flags)].copy()

                        m1, m2, m3 = st.columns(3)
                        m1.metric("Total Flag Occurrences", f"{len(filtered):,}")
                        m2.metric("Rows With Selected Flags", f"{filtered['row_idx'].nunique():,}")
                        m3.metric("Variables Impacted", f"{filtered['variable'].nunique():,}")

                        flag_counts = (
                            filtered['flag']
                            .value_counts()
                            .rename_axis("Flag")
                            .reset_index(name="Count")
                        )
                        total_rows_for_pct = max(1, len(df_viz))
                        total_occ_for_pct = max(1, len(filtered))
                        rows_with_flag = (
                            filtered.groupby('flag')['row_idx']
                            .nunique()
                            .rename("Rows_With_Flag")
                            .reset_index()
                            .rename(columns={"flag": "Flag"})
                        )
                        flag_counts = flag_counts.merge(rows_with_flag, on="Flag", how="left")
                        flag_counts["Rows_With_Flag"] = flag_counts["Rows_With_Flag"].fillna(0).astype(int)
                        flag_counts["Pct_of_Total_Rows"] = (
                            (flag_counts["Rows_With_Flag"] / total_rows_for_pct) * 100.0
                        ).round(2)
                        flag_counts["Pct_of_Selected_Flag_Occurrences"] = (
                            (flag_counts["Count"] / total_occ_for_pct) * 100.0
                        ).round(2)
                        variable_counts = (
                            filtered['variable']
                            .value_counts()
                            .rename_axis("Variable")
                            .reset_index(name="Count")
                        )

                        chart_col1, chart_col2 = st.columns(2)
                        with chart_col1:
                            st.subheader("Flag Frequency")
                            st.bar_chart(flag_counts.set_index("Flag")[["Count"]])
                            st.subheader("Flag Frequency (% of Total Rows)")
                            st.bar_chart(flag_counts.set_index("Flag")[["Pct_of_Total_Rows"]])
                            st.dataframe(flag_counts, use_container_width=True, hide_index=True)

                        with chart_col2:
                            st.subheader("Top Variables by Flag Count")
                            max_top = max(1, min(30, len(variable_counts)))
                            default_top = min(12, max_top)
                            top_n = st.slider(
                                "Number of variables",
                                min_value=1,
                                max_value=max_top,
                                value=default_top,
                                key="viz_top_variables"
                            )
                            st.bar_chart(variable_counts.head(top_n).set_index("Variable"))

                        st.subheader("Flag Count Matrix (Variable x Flag)")
                        matrix = (
                            filtered
                            .groupby(['variable', 'flag'])
                            .size()
                            .unstack(fill_value=0)
                            .sort_index()
                        )
                        st.dataframe(matrix, use_container_width=True)

                        freq_choice = "N/A"
                        total_over_time = pd.DataFrame()
                        by_flag_over_time = pd.DataFrame()

                        if 'TIMESTAMP' in df_viz.columns and df_viz['TIMESTAMP'].notna().any():
                            st.subheader("Flags Over Time")
                            freq_choice = st.selectbox(
                                "Time aggregation",
                                ["Daily", "Weekly", "Monthly"],
                                key="viz_time_freq"
                            )
                            freq_map = {"Daily": "D", "Weekly": "W", "Monthly": "M"}
                            total_over_time, by_flag_over_time = compute_flag_trend_tables(
                                filtered_flags=filtered,
                                df_viz=df_viz,
                                freq_code=freq_map[freq_choice]
                            )
                            if not total_over_time.empty:
                                st.line_chart(total_over_time)
                                st.area_chart(with_display_flag_columns(by_flag_over_time))
                            else:
                                st.info("No timestamped rows available for trend charts.")

                        ts_has_values = 'TIMESTAMP' in df_viz.columns and df_viz['TIMESTAMP'].notna().any()
                        if ts_has_values:
                            ts_start = df_viz['TIMESTAMP'].min().strftime("%Y-%m-%d")
                            ts_end = df_viz['TIMESTAMP'].max().strftime("%Y-%m-%d")
                        else:
                            ts_start = "N/A"
                            ts_end = "N/A"

                        summary_df = pd.DataFrame([
                            {"Metric": "Source File", "Value": selected_qc_file},
                            {"Metric": "Report Generated (UTC)", "Value": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")},
                            {"Metric": "Selected Flags", "Value": ", ".join(selected_flags)},
                            {"Metric": "Total Rows", "Value": len(df_viz)},
                            {"Metric": "Total Columns", "Value": df_viz.shape[1]},
                            {"Metric": "Date Range Start", "Value": ts_start},
                            {"Metric": "Date Range End", "Value": ts_end},
                            {"Metric": "Total Flag Occurrences", "Value": len(filtered)},
                            {"Metric": "Rows With Selected Flags", "Value": filtered['row_idx'].nunique()},
                            {"Metric": "Variables Impacted", "Value": filtered['variable'].nunique()},
                            {"Metric": "Time Aggregation", "Value": freq_choice},
                        ])

                        report_bytes = build_qc_viz_report_xlsx(
                            summary_df=summary_df,
                            flag_counts=flag_counts,
                            variable_counts=variable_counts,
                            matrix=matrix,
                            total_over_time=total_over_time,
                            by_flag_over_time=by_flag_over_time,
                        )

                        report_name = selected_qc_file.replace(".csv", "_visualization_report.xlsx")
                        st.download_button(
                            label="Download Visualization Report (Excel)",
                            data=report_bytes,
                            file_name=report_name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )

    # --- Tab 4: Save Trend Graphs ---
    with tab4:
        st.header("4. Save Trend Graphs")
        st.caption("Generate and save Daily/Weekly trend graph files from the final QA/QC output.")

        if os.path.exists(output_dir):
            qc_files_save = sorted([f for f in os.listdir(output_dir) if is_qaqc_output_file(f)])
        else:
            qc_files_save = []

        if not qc_files_save:
            st.info("No QA/QC output files found. Run Tab 2 first to generate a tidy QA/QC file.")
        else:
            selected_qc_file_save = st.selectbox(
                "Select QA/QC File",
                qc_files_save,
                key="save_graphs_file_select"
            )

            if selected_qc_file_save:
                qc_path_save = os.path.join(output_dir, selected_qc_file_save)
                with st.spinner("Loading QA/QC file..."):
                    df_save, flags_long_save = load_qc_visualization_data(qc_path_save)

                if flags_long_save.empty:
                    st.warning("No *_Flag values found in this file.")
                else:
                    all_flags_save = sorted(flags_long_save['flag'].unique().tolist())
                    default_flags_save = [f for f in all_flags_save if f != "P"] or all_flags_save

                    selected_flags_save = st.multiselect(
                        "Flags to include in saved graphs",
                        options=all_flags_save,
                        default=default_flags_save,
                        key="save_graphs_flags"
                    )
                    variable_options_save = (
                        flags_long_save['variable']
                        .value_counts()
                        .index
                        .tolist()
                    )
                    default_variable_count = min(5, len(variable_options_save))
                    selected_variables_save = st.multiselect(
                        "Variables for daily % by variable graph",
                        options=variable_options_save,
                        default=variable_options_save[:default_variable_count],
                        key="save_graphs_daily_pct_variables",
                        help="Daily percentages are computed per variable as: "
                             "(flag count for variable/day) / (records for variable/day). "
                             "One PNG is generated per selected variable."
                    )
                    freq_save = st.multiselect(
                        "Frequencies to save",
                        options=["Daily", "Weekly", "Monthly"],
                        default=["Daily", "Weekly"],
                        key="save_graphs_freq"
                    )
                    ts_available_save = (
                        'TIMESTAMP' in df_save.columns
                        and pd.to_datetime(df_save['TIMESTAMP'], errors='coerce').notna().any()
                    )
                    min_date_save = None
                    max_date_save = None
                    daily_range_enabled = False
                    daily_range_start = None
                    daily_range_end = None
                    if ts_available_save:
                        ts_dates_save = (
                            pd.to_datetime(df_save['TIMESTAMP'], errors='coerce')
                            .dropna()
                            .dt.date
                        )
                        if not ts_dates_save.empty:
                            min_date_save = ts_dates_save.min()
                            max_date_save = ts_dates_save.max()
                            daily_range_enabled = st.checkbox(
                                "Limit Daily outputs to date range",
                                value=False,
                                key="save_graphs_daily_range_enable",
                                help="Applies to Daily trend outputs and Daily % by variable outputs."
                            )
                            daily_range_values = st.date_input(
                                "Daily output date range",
                                value=(min_date_save, max_date_save),
                                min_value=min_date_save,
                                max_value=max_date_save,
                                key="save_graphs_daily_date_range"
                            )
                            if isinstance(daily_range_values, (tuple, list)):
                                if len(daily_range_values) >= 2:
                                    daily_range_start, daily_range_end = daily_range_values[0], daily_range_values[1]
                                elif len(daily_range_values) == 1:
                                    daily_range_start = daily_range_values[0]
                                    daily_range_end = daily_range_values[0]
                            else:
                                daily_range_start = daily_range_values
                                daily_range_end = daily_range_values
                            if daily_range_start and daily_range_end and daily_range_start > daily_range_end:
                                daily_range_start, daily_range_end = daily_range_end, daily_range_start

                    solar_overlay_options = [
                        c for c in ['SWalbedo_Avg', 'SWin_Avg', 'SWout_Avg']
                        if c in df_save.columns
                    ]
                    selected_solar_overlay_vars = st.multiselect(
                        "Solar variables for 15-min flag overlay",
                        options=solar_overlay_options,
                        default=solar_overlay_options,
                        key="save_graphs_solar_overlay_vars",
                        help="Plots 15-minute values with flagged points overlaid."
                    )
                    overlay_range_enabled = False
                    overlay_range_start = None
                    overlay_range_end = None
                    if min_date_save is not None and max_date_save is not None:
                        overlay_range_enabled = st.checkbox(
                            "Limit 15-min solar overlay to date range",
                            value=True,
                            key="save_graphs_overlay_range_enable",
                        )
                        overlay_range_values = st.date_input(
                            "15-min solar overlay date range",
                            value=(min_date_save, max_date_save),
                            min_value=min_date_save,
                            max_value=max_date_save,
                            key="save_graphs_overlay_date_range"
                        )
                        if isinstance(overlay_range_values, (tuple, list)):
                            if len(overlay_range_values) >= 2:
                                overlay_range_start, overlay_range_end = overlay_range_values[0], overlay_range_values[1]
                            elif len(overlay_range_values) == 1:
                                overlay_range_start = overlay_range_values[0]
                                overlay_range_end = overlay_range_values[0]
                        else:
                            overlay_range_start = overlay_range_values
                            overlay_range_end = overlay_range_values
                        if overlay_range_start and overlay_range_end and overlay_range_start > overlay_range_end:
                            overlay_range_start, overlay_range_end = overlay_range_end, overlay_range_start
                    default_save_dir = os.path.join(output_dir, "saved_graphs")
                    save_dir = st.text_input(
                        "Save directory",
                        value=default_save_dir,
                        key="save_graphs_dir"
                    )

                    if st.button("Generate and Save Graph Files", type="primary", key="save_graphs_btn"):
                        if not selected_flags_save:
                            st.warning("Select at least one flag.")
                        elif not selected_variables_save:
                            st.warning("Select at least one variable for the daily % graph.")
                        elif not freq_save:
                            st.warning("Select at least one frequency.")
                        else:
                            filtered_save = flags_long_save[
                                flags_long_save['flag'].isin(selected_flags_save)
                            ].copy()
                            daily_filtered_save = filtered_save
                            daily_range_label = "full range"
                            if (
                                daily_range_enabled
                                and ts_available_save
                                and daily_range_start is not None
                                and daily_range_end is not None
                            ):
                                ts_save = pd.to_datetime(df_save['TIMESTAMP'], errors='coerce')
                                start_ts = pd.Timestamp(daily_range_start)
                                end_ts = pd.Timestamp(daily_range_end) + timedelta(days=1) - timedelta(microseconds=1)
                                mask_daily_range = ts_save.between(start_ts, end_ts, inclusive='both')
                                daily_row_idx = set(df_save.index[mask_daily_range])
                                daily_filtered_save = filtered_save[
                                    filtered_save['row_idx'].isin(daily_row_idx)
                                ].copy()
                                daily_range_label = f"{daily_range_start} to {daily_range_end}"

                            if filtered_save.empty:
                                st.warning("No matching flag rows were found for the selected flags.")
                            else:
                                os.makedirs(save_dir, exist_ok=True)
                                created_paths = []
                                base_name = os.path.splitext(selected_qc_file_save)[0]
                                run_stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                                freq_map_save = {"Daily": "D", "Weekly": "W", "Monthly": "M"}

                                # Fresh percentage summary for this exact Save run.
                                flag_counts_save = (
                                    filtered_save['flag']
                                    .value_counts()
                                    .rename_axis("Flag")
                                    .reset_index(name="Count")
                                )
                                total_rows_for_pct_save = max(1, len(df_save))
                                total_occ_for_pct_save = max(1, len(filtered_save))
                                rows_with_flag_save = (
                                    filtered_save.groupby('flag')['row_idx']
                                    .nunique()
                                    .rename("Rows_With_Flag")
                                    .reset_index()
                                    .rename(columns={"flag": "Flag"})
                                )
                                flag_counts_save = flag_counts_save.merge(rows_with_flag_save, on="Flag", how="left")
                                flag_counts_save["Rows_With_Flag"] = flag_counts_save["Rows_With_Flag"].fillna(0).astype(int)
                                flag_counts_save["Pct_of_Total_Rows"] = (
                                    (flag_counts_save["Rows_With_Flag"] / total_rows_for_pct_save) * 100.0
                                ).round(2)
                                flag_counts_save["Pct_of_Selected_Flag_Occurrences"] = (
                                    (flag_counts_save["Count"] / total_occ_for_pct_save) * 100.0
                                ).round(2)

                                st.subheader("Flag Percentages (This Save Run)")
                                st.dataframe(flag_counts_save, use_container_width=True, hide_index=True)

                                pct_csv_path = os.path.join(
                                    save_dir,
                                    f"{base_name}_flag_percentages_{run_stamp}.csv"
                                )
                                pct_xlsx_path = os.path.join(
                                    save_dir,
                                    f"{base_name}_flag_percentages_{run_stamp}.xlsx"
                                )
                                pct_png_path = os.path.join(
                                    save_dir,
                                    f"{base_name}_flag_percentages_{run_stamp}.png"
                                )

                                flag_counts_save.to_csv(pct_csv_path, index=False)
                                with pd.ExcelWriter(pct_xlsx_path, engine="openpyxl") as writer:
                                    flag_counts_save.to_excel(writer, sheet_name="Flag Percentages", index=False)

                                pct_plot_df = flag_counts_save.sort_values("Pct_of_Total_Rows", ascending=False)
                                fig, ax = plt.subplots(figsize=(10, 4.8))
                                ax.bar(pct_plot_df["Flag"], pct_plot_df["Pct_of_Total_Rows"])
                                ax.set_ylabel("% of Total Rows")
                                ax.set_xlabel("Flag")
                                ax.set_title("Flag Frequency (% of Total Rows)")
                                fig.tight_layout()

                                pct_png_buffer = io.BytesIO()
                                fig.savefig(pct_png_buffer, format="png", dpi=180)
                                pct_png_buffer.seek(0)
                                pct_png_bytes = pct_png_buffer.getvalue()
                                plt.close(fig)

                                with open(pct_png_path, "wb") as f:
                                    f.write(pct_png_bytes)

                                st.image(pct_png_bytes, caption="Flag percentages (% of total rows)", use_container_width=True)
                                st.download_button(
                                    label="Download Flag Percentage Chart (PNG)",
                                    data=pct_png_bytes,
                                    file_name=os.path.basename(pct_png_path),
                                    mime="image/png",
                                    key=f"download_pct_png_{run_stamp}"
                                )

                                created_paths.extend([pct_csv_path, pct_xlsx_path, pct_png_path])

                                # Daily % by variable and flag.
                                daily_var_pct = compute_daily_variable_flag_percent_table(
                                    filtered_flags=daily_filtered_save,
                                    df_viz=df_save
                                )
                                if daily_var_pct.empty:
                                    if daily_range_enabled:
                                        st.info(f"No timestamped rows available for daily variable-percentage graph in range {daily_range_label}.")
                                    else:
                                        st.info("No timestamped rows available for daily variable-percentage graph.")
                                else:
                                    daily_var_pct = daily_var_pct[
                                        daily_var_pct['variable'].isin(selected_variables_save)
                                    ].copy()
                                    if daily_var_pct.empty:
                                        st.info("No rows available for the selected variable(s) in daily variable-percentage graph.")
                                    else:
                                        daily_var_pct_display = daily_var_pct[
                                            [
                                                "TIMESTAMP", "variable", "flag", "Flag_Count",
                                                "Variable_Daily_Count", "Pct_of_Variable_Daily_Count"
                                            ]
                                        ].copy()
                                        daily_var_pct_display.rename(
                                            columns={"flag": "Flag"},
                                            inplace=True
                                        )
                                        daily_var_pct_pivot_table = (
                                            daily_var_pct_display
                                            .pivot_table(
                                                index=["TIMESTAMP", "variable"],
                                                columns="Flag",
                                                values="Pct_of_Variable_Daily_Count",
                                                aggfunc="first",
                                                fill_value=0
                                            )
                                            .reset_index()
                                        )

                                        st.subheader("Daily Flag % by Variable")
                                        st.dataframe(
                                            daily_var_pct_display,
                                            use_container_width=True,
                                            hide_index=True
                                        )

                                        daily_pct_csv_path = os.path.join(
                                            save_dir,
                                            f"{base_name}_daily_variable_flag_percentages_{run_stamp}.csv"
                                        )
                                        daily_pct_xlsx_path = os.path.join(
                                            save_dir,
                                            f"{base_name}_daily_variable_flag_percentages_{run_stamp}.xlsx"
                                        )

                                        daily_var_pct_display.to_csv(daily_pct_csv_path, index=False)
                                        with pd.ExcelWriter(daily_pct_xlsx_path, engine="openpyxl") as writer:
                                            daily_var_pct_display.to_excel(
                                                writer,
                                                sheet_name="Daily Variable Flag %",
                                                index=False
                                            )
                                            daily_var_pct_pivot_table.to_excel(
                                                writer,
                                                sheet_name="Daily Variable Flag % Pivot",
                                                index=False
                                            )
                                        created_paths.extend([
                                            daily_pct_csv_path,
                                            daily_pct_xlsx_path
                                        ])

                                        png_paths = []
                                        for var_name in selected_variables_save:
                                            var_df = daily_var_pct[daily_var_pct["variable"] == var_name].copy()
                                            if var_df.empty:
                                                continue

                                            var_df["Flag_Label"] = var_df["flag"].apply(format_flag_label)
                                            var_pivot = (
                                                var_df
                                                .pivot_table(
                                                    index="TIMESTAMP",
                                                    columns="Flag_Label",
                                                    values="Pct_of_Variable_Daily_Count",
                                                    aggfunc="first",
                                                    fill_value=0
                                                )
                                                .sort_index()
                                            )
                                            if var_pivot.empty:
                                                continue

                                            var_slug = re.sub(r"[^A-Za-z0-9]+", "_", str(var_name)).strip("_") or "variable"
                                            var_png_path = os.path.join(
                                                save_dir,
                                                f"{base_name}_daily_variable_flag_percentages_{var_slug}_{run_stamp}.png"
                                            )
                                            var_png_bytes = build_daily_variable_flag_pct_png(
                                                var_pivot,
                                                f"{base_name} - {var_name} Daily Flag %"
                                            )

                                            with open(var_png_path, "wb") as f:
                                                f.write(var_png_bytes)

                                            st.subheader(f"Daily Flag %: {var_name}")
                                            st.image(
                                                var_png_bytes,
                                                caption=f"Daily flag percentages for {var_name}",
                                                use_container_width=True
                                            )
                                            st.download_button(
                                                label=f"Download {var_name} Daily % Chart (PNG)",
                                                data=var_png_bytes,
                                                file_name=os.path.basename(var_png_path),
                                                mime="image/png",
                                                key=f"download_daily_var_pct_png_{var_slug}_{run_stamp}"
                                            )
                                            png_paths.append(var_png_path)

                                        if not png_paths:
                                            st.info("No daily variable-percentage PNGs were generated for the selected variables.")
                                        created_paths.extend(png_paths)

                                # 15-min solar value + flagged-point overlays.
                                if selected_solar_overlay_vars:
                                    overlay_start_ts = None
                                    overlay_end_ts = None
                                    overlay_range_label = "full range"
                                    if (
                                        overlay_range_enabled
                                        and overlay_range_start is not None
                                        and overlay_range_end is not None
                                    ):
                                        overlay_start_ts = pd.Timestamp(overlay_range_start)
                                        overlay_end_ts = (
                                            pd.Timestamp(overlay_range_end)
                                            + timedelta(days=1)
                                            - timedelta(microseconds=1)
                                        )
                                        overlay_range_label = f"{overlay_range_start} to {overlay_range_end}"

                                    overlay_paths = []
                                    for solar_var in selected_solar_overlay_vars:
                                        solar_title = f"{base_name} - {solar_var} (15-min) with Flags"
                                        if overlay_range_enabled:
                                            solar_title = f"{solar_title} ({overlay_range_label})"
                                        solar_png_bytes = build_15min_variable_flag_overlay_png(
                                            df_viz=df_save,
                                            filtered_flags=filtered_save,
                                            variable=solar_var,
                                            title=solar_title,
                                            start_ts=overlay_start_ts,
                                            end_ts=overlay_end_ts
                                        )
                                        if solar_png_bytes is None:
                                            st.info(f"No plottable 15-min data for {solar_var} in selected range.")
                                            continue

                                        solar_slug = re.sub(r"[^A-Za-z0-9]+", "_", str(solar_var)).strip("_") or "solar"
                                        solar_png_path = os.path.join(
                                            save_dir,
                                            f"{base_name}_15min_flag_overlay_{solar_slug}_{run_stamp}.png"
                                        )
                                        with open(solar_png_path, "wb") as f:
                                            f.write(solar_png_bytes)

                                        st.subheader(f"15-min Solar Overlay: {solar_var}")
                                        st.image(
                                            solar_png_bytes,
                                            caption=f"{solar_var} with flagged points ({overlay_range_label})",
                                            use_container_width=True
                                        )
                                        st.download_button(
                                            label=f"Download {solar_var} 15-min Overlay (PNG)",
                                            data=solar_png_bytes,
                                            file_name=os.path.basename(solar_png_path),
                                            mime="image/png",
                                            key=f"download_15min_overlay_{solar_slug}_{run_stamp}"
                                        )
                                        overlay_paths.append(solar_png_path)

                                    if overlay_paths:
                                        created_paths.extend(overlay_paths)
                                    else:
                                        st.info("No 15-min solar overlay plots were generated.")

                                for freq_label in freq_save:
                                    freq_filtered_flags = daily_filtered_save if freq_label == "Daily" else filtered_save
                                    total_t, by_flag_t = compute_flag_trend_tables(
                                        filtered_flags=freq_filtered_flags,
                                        df_viz=df_save,
                                        freq_code=freq_map_save[freq_label]
                                    )
                                    if total_t.empty:
                                        if freq_label == "Daily" and daily_range_enabled:
                                            st.info(f"No timestamped rows available for {freq_label} trend in range {daily_range_label}.")
                                        else:
                                            st.info(f"No timestamped rows available for {freq_label} trend.")
                                        continue

                                    title = f"{base_name} - {freq_label} Flag Trend"
                                    if freq_label == "Daily" and daily_range_enabled:
                                        title = f"{base_name} - {freq_label} Flag Trend ({daily_range_label})"
                                    png_bytes = build_trend_png(total_t, by_flag_t, title)
                                    freq_key = freq_label.lower()

                                    png_path = os.path.join(
                                        save_dir,
                                        f"{base_name}_{freq_key}_trend_{run_stamp}.png"
                                    )
                                    total_csv_path = os.path.join(
                                        save_dir,
                                        f"{base_name}_{freq_key}_trend_total_{run_stamp}.csv"
                                    )
                                    by_flag_csv_path = os.path.join(
                                        save_dir,
                                        f"{base_name}_{freq_key}_trend_by_flag_{run_stamp}.csv"
                                    )

                                    with open(png_path, "wb") as f:
                                        f.write(png_bytes)
                                    total_t.reset_index().to_csv(total_csv_path, index=False)
                                    by_flag_t.reset_index().to_csv(by_flag_csv_path, index=False)

                                    created_paths.extend([png_path, total_csv_path, by_flag_csv_path])

                                    st.subheader(f"{freq_label} Output")
                                    st.image(png_bytes, caption=f"{freq_label} trend graph", use_container_width=True)
                                    st.download_button(
                                        label=f"Download {freq_label} Graph (PNG)",
                                        data=png_bytes,
                                        file_name=os.path.basename(png_path),
                                        mime="image/png",
                                        key=f"download_png_{freq_key}_{run_stamp}"
                                    )

                                if created_paths:
                                    st.success(f"Saved {len(created_paths)} files to: {save_dir}")
                                    st.write("Saved files:")
                                    for path in created_paths:
                                        st.code(path)

if __name__ == "__main__":
    main()
